import os
import threading
import re
import logging
import jdatetime
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
from supabase import create_client, Client
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, CallbackQueryHandler, filters
from telegram.request import HTTPXRequest
from io import BytesIO
from http.server import BaseHTTPRequestHandler, HTTPServer

TEHRAN_TZ = ZoneInfo("Asia/Tehran")
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

logger = logging.getLogger(__name__)
# ==========================================
# بارگذاری متغیرهای محیطی
# ==========================================
load_dotenv()

TOKEN = os.environ.get("BOT_TOKEN")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
ALLOWED_USER_ID = int(os.environ.get("ALLOWED_USER_ID", "0"))

if not TOKEN or not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("متغیرهای محیطی تنظیم نشده‌اند!")

if ALLOWED_USER_ID <= 0:
    raise ValueError("ALLOWED_USER_ID باید شناسه عددی تلگرام شما باشد (از @userinfobot بگیرید).")

# ==========================================
# اتصال به Supabase
# ==========================================
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


# ==========================================
# Health Server برای Render
# ==========================================


# ==========================================
# منوی اصلی
# ==========================================
MAIN_KEYBOARD = [
    ["📥 ثبت هزینه", "⚡️ هزینه‌های سریع"],
    ["📊 گزارش‌ها", "✏️ مدیریت هزینه‌ها"],
    ["⚙️ تنظیمات"],
]

def main_keyboard():
    return ReplyKeyboardMarkup(MAIN_KEYBOARD, resize_keyboard=True)

def back_keyboard():
    return ReplyKeyboardMarkup([["🔙 بازگشت"]], resize_keyboard=True)

# ==========================================
# دسته‌بندی‌های پیش‌فرض
# ==========================================
DEFAULT_CATEGORIES = [
    "🍔 غذا", "☕ کافه", "🚕 حمل‌ونقل", "🛒 خرید", "🏠 خانه",
    "🎮 تفریح", "💊 درمان", "💳 قبض", "📦 سایر"
]

# ==========================================
# کلمات کلیدی برای تشخیص خودکار دسته‌بندی
# ==========================================
DEFAULT_CATEGORY_KEYWORDS = {
    "🍔 غذا": ["ناهار", "شام", "صبحانه", "غذا", "پیتزا", "برگر", "کباب", "ساندویچ", "فست‌فود", "رستوران"],
    "🚕 حمل‌ونقل": ["تاکسی", "اتوبوس", "مترو", "اسنپ", "قطار", "بی‌آرتی", "پمپ بنزین"],
    "🛒 خرید": ["خرید", "فروشگاه", "سوپرمارکت", "لباس", "کفش", "لوازم", "خواربار"],
    "🏠 خانه": ["خانه", "اجاره", "تعمیرات", "مبلمان", "لوازم خانگی"],
    "🎮 تفریح": ["تفریح", "سینما", "بازی", "کنسرت", "ورزش", "باشگاه"],
    "💊 درمان": ["درمان", "دارو", "دکتر", "بیمارستان", "آزمایشگاه", "مطب"],
    "💳 قبض": ["قبض برق", "برق", "قبض گاز", "قبض اب", "تلفن", "اینترنت", "موبایل"],
    "☕ کافه": ["کافه", "قهوه", "چای", "نسکافه", "کاپوچینو", "ابمیوه"],
}

def init_db():
    """بررسی اتصال به Supabase و ساخت دسته‌ها/کلمات کلیدی پیش‌فرض (بدون تکرار)."""
    try:
        supabase.table("expenses").select("*").limit(1).execute()
    except Exception as e:
        logger.exception(f"خطا در بررسی دیتابیس: {e}")
        print("⚠️ لطفاً جدول‌ها را در Supabase بسازید!")
        return

    # دسته‌بندی‌های پیش‌فرض (فقط در صورت نبودن، اضافه می‌شوند)
    category_ids = {}
    for category in DEFAULT_CATEGORIES:
        try:
            existing = supabase.table("categories").select("id").eq("name", category).execute()
            if existing.data:
                category_ids[category] = existing.data[0]["id"]
                continue
            inserted = supabase.table("categories").insert({"name": category}).execute()
            if inserted.data:
                category_ids[category] = inserted.data[0]["id"]
        except Exception as e:
            logger.warning(f"دسته پیش‌فرض «{category}» اضافه نشد: {e}")

    # کلمات کلیدی پیش‌فرض (فقط کلماتی که هنوز ثبت نشده‌اند اضافه می‌شوند)
    for category_name, keywords in DEFAULT_CATEGORY_KEYWORDS.items():
        cat_id = category_ids.get(category_name)
        if not cat_id:
            continue
        try:
            existing = supabase.table("category_keywords").select("keyword").eq("category_id", cat_id).execute()
            existing_keywords = {
                (row.get("keyword") or "").strip().casefold()
                for row in existing.data or []
            }
            for keyword in keywords:
                if keyword.strip().casefold() in existing_keywords:
                    continue
                supabase.table("category_keywords").insert({"keyword": keyword, "category_id": cat_id}).execute()
        except Exception as e:
            logger.warning(f"کلمات پیش‌فرض برای «{category_name}» اضافه نشد: {e}")

# ==========================================
# توابع دیتابیس
# ==========================================
def get_categories():
    response = supabase.table("categories").select("*").execute()

    desired_order = [
        "🍔 غذا",
        "☕ کافه",
        "🚕 حمل‌ونقل",
        "🛒 خرید",
        "🏠 خانه",
        "🎮 تفریح",
        "💊 درمان",
        "💳 قبض",
        "📦 سایر",
    ]

    categories = [(row["id"], row["name"]) for row in response.data]

    order_map = {name: i for i, name in enumerate(desired_order)}

    categories.sort(
        key=lambda item: order_map.get(item[1], len(desired_order))
    )

    return categories

# ==========================================
# گزارش هزینه بر اساس دسته‌بندی و بازه زمانی
# ==========================================
def get_category_report_expenses(
    user_id,
    category_name,
    start_date,
    end_date
):
    """
    دریافت هزینه‌های یک دسته‌بندی در یک بازه زمانی.
    start_date و end_date با فرمت YYYY-MM-DD هستند.
    """

    response = (
        supabase
        .table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .eq("category", category_name)
        .gte("created_at", f"{start_date} 00:00:00")
        .lte("created_at", f"{end_date} 23:59:59")
        .order("created_at", desc=True)
        .execute()
    )

    rows = response.data or []

    return [
        (
            row["id"],
            row["amount"],
            row["description"],
            row["category"],
            row["created_at"]
        )
        for row in rows
    ]
def category_exists(name):
    response = supabase.table("categories").select("id").eq("name", name).execute()
    return len(response.data) > 0

def add_category(name):
    try:
        supabase.table("categories").insert({"name": name}).execute()
        return True
    except Exception as e:
        logger.exception(f"خطا در افزودن دسته: {e}")
        return False

def rename_category(category_id, new_name):
    response = supabase.table("categories").select("name").eq("id", category_id).execute()
    if not response.data:
        return False
    old_name = response.data[0]["name"]
    try:
        supabase.table("categories").update({"name": new_name}).eq("id", category_id).execute()
        supabase.table("expenses").update({"category": new_name}).eq("category", old_name).execute()
        return True
    except Exception as e:
        logger.exception(f"خطا در تغییر نام دسته: {e}")
        return False

def delete_category(category_id):
    response = supabase.table("categories").select("name").eq("id", category_id).execute()
    if not response.data:
        return False
    category_name = response.data[0]["name"]
    if category_name == "📦 سایر":
        return False
    supabase.table("expenses").update({"category": "📦 سایر"}).eq("category", category_name).execute()
    supabase.table("categories").delete().eq("id", category_id).execute()
    return True
    
# ==========================================
# توابع دیتابیس برای هزینه‌های سریع
# ==========================================
def add_quick_expense(user_id, name, amount, category):
    """افزودن هزینه سریع جدید"""
    data = {
        "user_id": user_id,
        "name": name,
        "amount": amount,
        "category": category,
        "created_at": datetime.now(TEHRAN_TZ).isoformat()
    }
    supabase.table("quick_expenses").insert(data).execute()

def get_quick_expenses(user_id):
    """دریافت لیست هزینه‌های سریع کاربر"""
    response = supabase.table("quick_expenses").select("*").eq("user_id", user_id).order("id").execute()
    return response.data

def delete_quick_expense(user_id, expense_id):
    """حذف هزینه سریع"""
    response = supabase.table("quick_expenses").delete().eq("id", expense_id).eq("user_id", user_id).execute()
    return len(response.data) > 0

def update_quick_expense(user_id, expense_id, name, amount, category):
    """ویرایش هزینه سریع"""
    data = {"name": name, "amount": amount, "category": category}
    response = supabase.table("quick_expenses").update(data).eq("id", expense_id).eq("user_id", user_id).execute()
    return len(response.data) > 0    

def add_expense(user_id, amount, description, category):
    data = {
        "user_id": user_id,
        "amount": amount,
        "description": description,
        "category": category,
        "created_at": datetime.now(TEHRAN_TZ).isoformat()
    }
    supabase.table("expenses").insert(data).execute()

def get_expense(user_id, expense_id):
    response = supabase.table("expenses").select("*").eq("id", expense_id).eq("user_id", user_id).execute()
    if response.data:
        row = response.data[0]
        return (row["id"], row["amount"], row["description"], row["category"], row["created_at"])
    return None

def delete_expense(user_id, expense_id):
    response = supabase.table("expenses").delete().eq("id", expense_id).eq("user_id", user_id).execute()
    return len(response.data) > 0

def update_expense(user_id, expense_id, amount, description, category):
    data = {"amount": amount, "description": description, "category": category}
    response = supabase.table("expenses").update(data).eq("id", expense_id).eq("user_id", user_id).execute()
    return len(response.data) > 0

def get_recent_expenses(user_id, limit=10):
    response = supabase.table("expenses").select("*").eq("user_id", user_id).order("id", desc=True).limit(limit).execute()
    return [(row["id"], row["amount"], row["description"], row["category"], row["created_at"]) for row in response.data]

def get_day_expenses(user_id, date_text):
    # بازه‌ی دقیق یک روز؛ مستقل از طول زمان/میلی‌ثانیه‌ی created_at
    next_date = (datetime.strptime(date_text, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    response = (
        supabase.table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .gte("created_at", f"{date_text} 00:00:00")
        .lt("created_at", f"{next_date} 00:00:00")
        .order("id", desc=True)
        .execute()
    )
    return [(row["id"], row["amount"], row["description"], row["category"], row["created_at"]) for row in response.data]

def get_month_expenses(user_id, month_text):
    # ماه را با اولین روز ماه بعد محدود می‌کنیم؛ بنابراین برای فوریه، آوریل و ... هم درست است.
    first_day = datetime.strptime(f"{month_text}-01", "%Y-%m-%d")
    if first_day.month == 12:
        next_month = first_day.replace(year=first_day.year + 1, month=1, day=1)
    else:
        next_month = first_day.replace(month=first_day.month + 1, day=1)
    next_month_text = next_month.strftime("%Y-%m-%d")

    response = (
        supabase.table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .gte("created_at", f"{month_text}-01 00:00:00")
        .lt("created_at", f"{next_month_text} 00:00:00")
        .order("id", desc=True)
        .execute()
    )
    return response.data

def get_advanced_stats(user_id, start_date, end_date):
    response = supabase.table("expenses").select("*").eq("user_id", user_id).gte("created_at", f"{start_date} 00:00:00").lte("created_at", f"{end_date} 23:59:59").execute()
    rows = response.data
    if not rows:
        return (0, 0, 0, 0), [], []
    total = sum(r["amount"] for r in rows)
    count = len(rows)
    average = total // count if count > 0 else 0
    maximum = max(r["amount"] for r in rows) if rows else 0
    daily = {}
    category = {}
    for row in rows:
        date_key = row["created_at"][:10]
        daily[date_key] = daily.get(date_key, 0) + row["amount"]
        cat = row["category"]
        if cat not in category:
            category[cat] = {"total": 0, "count": 0}
        category[cat]["total"] += row["amount"]
        category[cat]["count"] += 1
    daily_rows = [(d, a, 1) for d, a in daily.items()]
    category_rows = [(c, d["total"], d["count"]) for c, d in category.items()]
    return (total, count, average, maximum), daily_rows, category_rows

# ==========================================
# توابع کمکی
# ==========================================
def normalize_digits(text):
    translation = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
    return text.translate(translation)

def parse_amount(text):
    text = normalize_digits(text)
    text = text.replace(",", "").replace("٬", "").replace("،", "").replace(" ", "")
    if not text.isdigit():
        return None
    amount = int(text)
    return amount if amount > 0 else None

def parse_expense_text(message):
    message = normalize_digits(message.strip())
    if not message:
        return None
    match = re.match(r"^([\d,\u066C\u060C]+)\s+(.+)$", message)
    if not match:
        return None
    amount_text = match.group(1)
    description = match.group(2).strip()
    amount = parse_amount(amount_text)
    if amount is None or not description:
        return None
    return amount, description

def detect_category(description):
    """تشخیص خودکار دسته‌بندی بر اساس کلمات ذخیره‌شده در دیتابیس"""

    description_lower = description.lower().strip()

    try:
        # دریافت کلمات کلیدی از دیتابیس
        response = (
            supabase
            .table("category_keywords")
            .select("keyword, category_id")
            .execute()
        )

        # دریافت دسته‌بندی‌ها و ساختن نقشه ID → نام
        categories = dict(get_categories())

        # بررسی کلمات کلیدی
        for row in response.data:
            keyword = (row.get("keyword") or "").strip().lower()
            category_id = row.get("category_id")

            if not keyword or not category_id:
                continue

            if keyword in description_lower:
                category = categories.get(category_id)

                if category:
                    return category

    except Exception as e:
        logger.error(f"خطا در تشخیص دسته‌بندی: {e}")

    # اگر هیچ کلمه‌ای پیدا نشد
    return "📦 سایر"
# ==========================================
# تابع تبدیل تاریخ میلادی به شمسی
# ==========================================
def to_jalali(date_str):
    """تبدیل تاریخ میلادی به شمسی با فرمت YYYY-MM-DD"""
    if not date_str:
        return ""
    try:
        # استخراج بخش تاریخ (10 کاراکتر اول)
        date_part = date_str[:10]
        year, month, day = map(int, date_part.split('-'))
        gregorian = datetime(year, month, day)
        jalali = jdatetime.date.fromgregorian(date=gregorian)
        return f"{jalali.year:04d}-{jalali.month:02d}-{jalali.day:02d}"
    except:
        # اگر خطایی رخ داد، همان تاریخ میلادی را برگردان
        return date_str[:10] if len(date_str) >= 10 else ""


def parse_date_input(date_text):
    """
    تنها موتور تشخیص تاریخ در کل ربات.

    ورودی:
        1405-05-23
        1405/05/23
        1405.05.23
        ۲۰۲۶-۰۸-۱۴
        2026-08-14

    خروجی:
        تاریخ میلادی با فرمت YYYY-MM-DD
        یا None
    """

    if not date_text:
        return None

    date_text = normalize_digits(str(date_text).strip())

    # یکسان‌سازی جداکننده‌ها
    date_text = (
        date_text
        .replace("/", "-")
        .replace(".", "-")
        .replace("\\", "-")
    )

    # حذف فاصله اطراف -
    date_text = re.sub(r"\s*-\s*", "-", date_text)

    parts = date_text.split("-")

    if len(parts) != 3:
        return None

    try:
        year, month, day = map(int, parts)
    except ValueError:
        return None

    try:
        # میلادی
        if 1700 <= year <= 3000:
            gregorian = datetime(year, month, day)
            return gregorian.strftime("%Y-%m-%d")

        # شمسی
        if 1200 <= year <= 1600:
            jalali = jdatetime.date(year, month, day)
            gregorian = jalali.togregorian()
            return gregorian.strftime("%Y-%m-%d")

    except (ValueError, TypeError):
        return None

    return None

def get_date_info(date_text):
    """
    اطلاعات کامل تاریخ را بر اساس parse_date_input برمی‌گرداند.

    تمام بخش‌های ربات باید برای تشخیص تاریخ
    از parse_date_input استفاده کنند.
    """

    if not date_text:
        return None

    original = normalize_digits(str(date_text).strip())

    gregorian = parse_date_input(original)

    if not gregorian:
        return None

    try:
        gregorian_date = datetime.strptime(
            gregorian,
            "%Y-%m-%d"
        )

        jalali_date = jdatetime.date.fromgregorian(
            date=gregorian_date
        )

        jalali = (
            f"{jalali_date.year:04d}-"
            f"{jalali_date.month:02d}-"
            f"{jalali_date.day:02d}"
        )

        # تشخیص نوع تقویم فقط برای اطلاعات خروجی
        normalized = (
            original
            .replace("/", "-")
            .replace(".", "-")
            .replace("\\", "-")
        )

        input_year = int(normalized.split("-")[0])

        calendar = (
            "gregorian"
            if 1700 <= input_year <= 3000
            else "jalali"
        )

        return {
            "input": original,
            "calendar": calendar,
            "gregorian": gregorian,
            "jalali": jalali
        }

    except (ValueError, TypeError):
        return None

def category_keyboard():
    categories = get_categories()
    keyboard = []
    row = []
    for _, name in categories:
        row.append(name)
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    keyboard.append(["🔙 بازگشت"])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def is_allowed(user_id):
    return user_id == ALLOWED_USER_ID

# ==========================================
# هندلرها
# ==========================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_allowed(user_id):
        await update.message.reply_text("⛔ شما اجازه استفاده از این ربات را ندارید.")
        return
    context.user_data.clear()
    await update.message.reply_text(
        "سلام 👋\n\n💰 دفتر هزینه شخصی آماده است.\n\n⚡ ثبت سریع:\n85 ناهار\n85000 خرید\n\nاز منوی پایین انتخاب کن.",
        reply_markup=main_keyboard()
    )

async def go_back(update, context):
    context.user_data.clear()
    await update.message.reply_text("🏠 برگشتیم به منوی اصلی.", reply_markup=main_keyboard())

async def render_quick_expenses_menu(update, context):
    """نمایش منوی هزینه‌های سریع (ارسال پیام جدید یا ویرایش پیام فعلی)"""
    user_id = update.effective_user.id
    quick_items = get_quick_expenses(user_id)

    keyboard = []
    row = []
    for item in quick_items:
        row.append(InlineKeyboardButton(
            f"{item['name']} ({item['amount']:,})",
            callback_data=f"quick_{item['id']}"
        ))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)

    keyboard.append([InlineKeyboardButton("⚙️ مدیریت هزینه‌های سریع", callback_data="quick_manage")])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="back_menu")])

    if quick_items:
        text = "🧾 **هزینه‌های سریع**\n\n"
        text += "یکی از گزینه‌های زیر رو انتخاب کن تا هزینه ثبت بشه:\n\n"
        text += "\n".join(
            f"• {item['name']} — {item['amount']:,} تومان — {item['category']}"
            for item in quick_items
        )
    else:
        text = (
            "🧾 **هزینه‌های سریع**\n\n"
            "هنوز هیچ هزینه سریعی ثبت نشده است.\n\n"
            "از «⚙️ مدیریت هزینه‌های سریع» می‌تونی "
            "یک مورد جدید اضافه کنی."
        )

    markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)


async def quick_expenses_menu(update, context):
    """منوی هزینه‌های سریع (از پیام متنی)"""
    user_id = update.effective_user.id

    if not is_allowed(user_id):
        return

    await render_quick_expenses_menu(update, context)


async def quick_menu_callback(update, context):
    """بازگشت به منوی هزینه‌های سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    await render_quick_expenses_menu(update, context)

async def quick_manage_callback(update, context):
    """منوی مدیریت هزینه‌های سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    buttons = [
        [InlineKeyboardButton("➕ افزودن هزینه سریع", callback_data="quick_add")],
        [InlineKeyboardButton("✏️ ویرایش هزینه سریع", callback_data="quick_edit")],
        [InlineKeyboardButton("🗑️ حذف هزینه سریع", callback_data="quick_delete")],
        [InlineKeyboardButton("🔙 بازگشت به هزینه‌های سریع", callback_data="quick_menu")],
    ]

    await query.edit_message_text(
        "⚙️ **مدیریت هزینه‌های سریع**\n\n"
        "می‌توانی هزینه‌های سریع رو مدیریت کنی:\n"
        "➕ افزودن هزینه جدید\n"
        "✏️ ویرایش مبلغ یا نام\n"
        "🗑️ حذف هزینه‌های غیرضروری\n\n"
        "یکی از گزینه‌های زیر رو انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def quick_callback(update, context):
    """ثبت هزینه‌های سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    # دریافت id هزینه سریع
    quick_id = int(query.data.replace("quick_", ""))

    # دریافت از دیتابیس
    response = supabase.table("quick_expenses").select("*").eq("id", quick_id).eq("user_id", user_id).execute()
    if not response.data:
        await query.edit_message_text("❌ هزینه سریع پیدا نشد.")
        return

    item = response.data[0]
    name = item["name"]
    amount = item["amount"]
    category = item["category"]

    # ثبت هزینه اصلی
    add_expense(user_id, amount, name, category)

    context.user_data.clear()

    await query.edit_message_text(
        f"✅ هزینه ثبت شد!\n\n"
        f"{category}\n"
        f"💰 {amount:,} تومان\n"
        f"📝 {name}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔙 بازگشت به منو", callback_data="back_menu")]
        ])
    )

async def expense_button(update, context):
    context.user_data.clear()
    context.user_data["waiting_for_expense"] = True
    await update.message.reply_text("➕ ثبت هزینه\n\nدسته‌بندی را انتخاب کن:", reply_markup=category_keyboard())

async def choose_category(update, context, category):
    context.user_data["selected_category"] = category
    context.user_data["waiting_for_expense"] = False
    context.user_data["waiting_for_amount"] = True
    await update.message.reply_text(
        f"{category}\n\nمبلغ و توضیح هزینه را وارد کن.\n\nمثال:\n85000 ناهار",
        reply_markup=back_keyboard()
    )

async def advanced_quick_callback(update, context):
    """دکمه‌های میانبر برای گزارش پیشرفته"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    today = datetime.now(TEHRAN_TZ).date()
    action = query.data

    if action == "adv_today":
        start_date = today.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    elif action == "adv_this_week":
    # هفته جاری: شنبه تا امروز
        days_since_saturday = (today.weekday() + 2) % 7

        start_of_week = today - timedelta(days=days_since_saturday)

        start_date = start_of_week.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    elif action == "adv_week":
        # هفته گذشته: شنبه تا جمعه
        days_since_saturday = (today.weekday() + 2) % 7

        start_of_this_week = today - timedelta(days=days_since_saturday)
        start_of_last_week = start_of_this_week - timedelta(days=7)
        end_of_last_week = start_of_this_week - timedelta(days=1)

        start_date = start_of_last_week.strftime("%Y-%m-%d")
        end_date = end_of_last_week.strftime("%Y-%m-%d")

    elif action == "adv_month":
        # اولین روز ماه شمسی
        today_jalali = jdatetime.date.fromgregorian(date=today)

        first_day_jalali = jdatetime.date(
            today_jalali.year,
            today_jalali.month,
            1
        )

        start_date = (
            first_day_jalali
            .togregorian()
            .strftime("%Y-%m-%d")
        )

        end_date = today.strftime("%Y-%m-%d")

    elif action == "adv_quarter":
        start_date = (today - timedelta(days=90)).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    elif action == "adv_manual":
        await query.edit_message_text(
            "📅 تاریخ شروع را وارد کن:\n\n"
            "مثال:\n"
            "2026-08-01"
        )

        context.user_data["waiting_advanced_start"] = True
        return

    else:
        await query.edit_message_text("❌ گزینه نامعتبر.")
        return

    context.user_data.clear()

    await show_advanced_report(
        update,
        context,
        start_date,
        end_date,
        from_callback=True
    )

async def show_advanced_report(update, context, start_date, end_date, from_callback=False):
    """نمایش گزارش پیشرفته با تاریخ شمسی + لیست هزینه‌ها با صفحه‌بندی ۵تایی"""
    user_id = update.effective_user.id

    # تبدیل تاریخ‌ها به شمسی برای نمایش
    start_jalali = to_jalali(start_date)
    end_jalali = to_jalali(end_date)

    # دریافت آمار
    (total, count, average, maximum), daily_rows, category_rows = get_advanced_stats(
        user_id,
        start_date,
        end_date
    )

    # دریافت تمام هزینه‌های همین بازه
    response = (
        supabase
        .table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .gte("created_at", f"{start_date} 00:00:00")
        .lte("created_at", f"{end_date} 23:59:59")
        .order("id", desc=True)
        .execute()
    )

    expense_rows = response.data or []

    if count == 0 or not expense_rows:
        if from_callback and update.callback_query:
            await update.callback_query.edit_message_text(
                "📊 در این بازه هزینه‌ای ثبت نشده."
            )
        else:
            await update.message.reply_text(
                "📊 در این بازه هزینه‌ای ثبت نشده.",
                reply_markup=main_keyboard()
            )

        context.user_data.clear()
        return

    # ذخیره اطلاعات گزارش برای صفحه‌بندی
    context.user_data["advanced_report_expenses"] = expense_rows
    context.user_data["advanced_report_start_date"] = start_date
    context.user_data["advanced_report_end_date"] = end_date
    context.user_data["advanced_report_page"] = 0

    # مرتب‌سازی دسته‌بندی‌ها بر اساس مبلغ
    category_rows_sorted = sorted(
        category_rows,
        key=lambda x: x[1],
        reverse=True
    )

    # عنوان هوشمند گزارش
    today = datetime.now(TEHRAN_TZ).strftime("%Y-%m-%d")

    if start_date == today and end_date == today:
        report_title = "📊 گزارش امروز"
    else:
        report_title = "📅 گزارش بر اساس تاریخ"

    # ساخت بخش ثابت گزارش
    text = f"{report_title}\n\n"
    text += f"📅 از {start_jalali}\n"
    text += f"📅 تا {end_jalali}\n\n"

    # آمار کلی
    text += "━━━━━━━━━━━━\n"
    text += f"💵 مجموع: {total:,} تومان\n"
    text += f"🧾 تعداد: {count}\n"
    text += f"📊 میانگین هر هزینه: {average:,}\n"
    text += f"🔝 بیشترین هزینه: {maximum:,}\n\n"

    # بر اساس دسته‌بندی
    text += "━━━━━━━━━━━━\n"
    text += "📊 بر اساس دسته‌بندی\n\n"

    for category, amount, cnt in category_rows_sorted:
        text += f"{category}\n"
        text += f"💰 {amount:,} تومان ({cnt} مورد)\n\n"

    # روند روزانه
    if daily_rows:
        text += "━━━━━━━━━━━━\n"
        text += "📅 روند روزانه\n\n"

        daily_rows_sorted = sorted(
            daily_rows,
            key=lambda x: x[0],
            reverse=True
        )

        for date_text, amount, _ in daily_rows_sorted:
            date_jalali = to_jalali(date_text)
            text += f"{date_jalali}: {amount:,} تومان\n"

    # نمایش صفحه اول لیست هزینه‌ها
    text, buttons = build_advanced_report_page(
        text,
        expense_rows,
        page=0
    )

    if from_callback and update.callback_query:
        await update.callback_query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )
    else:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )
async def show_advanced_report(update, context, start_date, end_date, from_callback=False):
    """نمایش گزارش پیشرفته با تاریخ شمسی + لیست هزینه‌ها با صفحه‌بندی ۵تایی"""
    user_id = update.effective_user.id

    # تبدیل تاریخ‌ها به شمسی برای نمایش
    start_jalali = to_jalali(start_date)
    end_jalali = to_jalali(end_date)

    # دریافت آمار
    (total, count, average, maximum), daily_rows, category_rows = get_advanced_stats(
        user_id,
        start_date,
        end_date
    )

    # دریافت تمام هزینه‌های همین بازه
    response = (
        supabase
        .table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .gte("created_at", f"{start_date} 00:00:00")
        .lte("created_at", f"{end_date} 23:59:59")
        .order("id", desc=True)
        .execute()
    )

    expense_rows = response.data or []

    if count == 0 or not expense_rows:
        if from_callback and update.callback_query:
            await update.callback_query.edit_message_text(
                "📊 در این بازه هزینه‌ای ثبت نشده."
            )
        else:
            await update.message.reply_text(
                "📊 در این بازه هزینه‌ای ثبت نشده.",
                reply_markup=main_keyboard()
            )

        context.user_data.clear()
        return

    # ذخیره اطلاعات گزارش برای صفحه‌بندی
    context.user_data["advanced_report_expenses"] = expense_rows
    context.user_data["advanced_report_start_date"] = start_date
    context.user_data["advanced_report_end_date"] = end_date
    context.user_data["advanced_report_page"] = 0

    # مرتب‌سازی دسته‌بندی‌ها بر اساس مبلغ
    category_rows_sorted = sorted(
        category_rows,
        key=lambda x: x[1],
        reverse=True
    )

    # عنوان هوشمند گزارش
    today = datetime.now(TEHRAN_TZ).strftime("%Y-%m-%d")

    if start_date == today and end_date == today:
        report_title = "📊 گزارش امروز"
    else:
        report_title = "📅 گزارش بر اساس تاریخ"

    # ساخت بخش ثابت گزارش
    text = f"{report_title}\n\n"
    text += f"📅 از {start_jalali}\n"
    text += f"📅 تا {end_jalali}\n\n"

    # آمار کلی
    text += "━━━━━━━━━━━━\n"
    text += f"💵 مجموع: {total:,} تومان\n"
    text += f"🧾 تعداد: {count}\n"
    text += f"📊 میانگین هر هزینه: {average:,}\n"
    text += f"🔝 بیشترین هزینه: {maximum:,}\n\n"

    # بر اساس دسته‌بندی
    text += "━━━━━━━━━━━━\n"
    text += "📊 بر اساس دسته‌بندی\n\n"

    for category, amount, cnt in category_rows_sorted:
        text += f"{category}\n"
        text += f"💰 {amount:,} تومان ({cnt} مورد)\n\n"

    # روند روزانه
    if daily_rows:
        text += "━━━━━━━━━━━━\n"
        text += "📅 روند روزانه\n\n"

        daily_rows_sorted = sorted(
            daily_rows,
            key=lambda x: x[0],
            reverse=True
        )

        for date_text, amount, _ in daily_rows_sorted:
            date_jalali = to_jalali(date_text)
            text += f"{date_jalali}: {amount:,} تومان\n"

    # نمایش صفحه اول لیست هزینه‌ها
    text, buttons = build_advanced_report_page(
        text,
        expense_rows,
        page=0
    )

    if from_callback and update.callback_query:
        await update.callback_query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )
    else:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )
async def advanced_report_page_callback(update, context):
    """جابجایی بین صفحات لیست هزینه‌های گزارش پیشرفته"""

    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_allowed(user_id):
        return

    # اطلاعات گزارش ذخیره شده؟
    expense_rows = context.user_data.get(
        "advanced_report_expenses"
    )

    start_date = context.user_data.get(
        "advanced_report_start_date"
    )

    end_date = context.user_data.get(
        "advanced_report_end_date"
    )

    if not expense_rows or not start_date or not end_date:
        await query.answer(
            "❌ اطلاعات این گزارش دیگر در دسترس نیست.",
            show_alert=True
        )
        return

    # صفحه درخواستی
    try:
        page = int(
            query.data.split(":")[1]
        )
    except (IndexError, ValueError):
        page = 0

    # دوباره آمار را دریافت می‌کنیم
    (
        total,
        count,
        average,
        maximum
    ), daily_rows, category_rows = get_advanced_stats(
        user_id,
        start_date,
        end_date
    )

    # مرتب‌سازی دسته‌بندی‌ها
    category_rows_sorted = sorted(
        category_rows,
        key=lambda x: x[1],
        reverse=True
    )

    start_jalali = to_jalali(start_date)
    end_jalali = to_jalali(end_date)

    today = datetime.now(
        TEHRAN_TZ
    ).strftime("%Y-%m-%d")

    if start_date == today and end_date == today:
        report_title = "📊 گزارش امروز"
    else:
        report_title = "📅 گزارش بر اساس تاریخ"

    # ساخت گزارش ثابت
    text = f"{report_title}\n\n"
    text += f"📅 از {start_jalali}\n"
    text += f"📅 تا {end_jalali}\n\n"

    text += "━━━━━━━━━━━━\n"
    text += f"💵 مجموع: {total:,} تومان\n"
    text += f"🧾 تعداد: {count}\n"
    text += f"📊 میانگین هر هزینه: {average:,}\n"
    text += f"🔝 بیشترین هزینه: {maximum:,}\n\n"

    # دسته‌بندی‌ها
    text += "━━━━━━━━━━━━\n"
    text += "📊 بر اساس دسته‌بندی\n\n"

    for category, amount, cnt in category_rows_sorted:
        text += f"{category}\n"
        text += f"💰 {amount:,} تومان ({cnt} مورد)\n\n"

    # روند روزانه
    if daily_rows:
        text += "━━━━━━━━━━━━\n"
        text += "📅 روند روزانه\n\n"

        daily_rows_sorted = sorted(
            daily_rows,
            key=lambda x: x[0],
            reverse=True
        )

        for date_text, amount, _ in daily_rows_sorted:
            date_jalali = to_jalali(date_text)
            text += f"{date_jalali}: {amount:,} تومان\n"

    # ساخت صفحه موردنظر
    text, buttons = build_advanced_report_page(
        text,
        expense_rows,
        page=page
    )

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup(buttons)
    )
# ==========================================
# گزارش بر اساس دسته‌بندی
# ==========================================

def category_report_period_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📅 امروز", callback_data="cat_report_today"),
            InlineKeyboardButton("📅 این هفته", callback_data="cat_report_this_week"),
        ],
        [
            InlineKeyboardButton("📅 هفته گذشته", callback_data="cat_report_last_week"),
            InlineKeyboardButton("📅 این ماه", callback_data="cat_report_this_month"),
        ],
        [
            InlineKeyboardButton("📅 سه ماه اخیر", callback_data="cat_report_quarter"),
            InlineKeyboardButton("✏️ بازه دلخواه", callback_data="cat_report_manual"),
        ],
        [InlineKeyboardButton("🔙 انتخاب دسته", callback_data="cat_report_back_category")],
        [InlineKeyboardButton("🔙 بازگشت به گزارش‌ها", callback_data="reports_menu")],
    ])


async def category_report_callback(update, context):
    """مدیریت گزارش بر اساس دسته‌بندی"""

    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_allowed(user_id):
        return

    action = query.data
        # ==========================================
    # بازگشت به انتخاب بازه زمانی
    # ==========================================
    if action == "cat_report_period_back":

        category_name = context.user_data.get(
            "category_report_category"
        )

        if not category_name:
            await category_report_show_categories(
                update,
                context,
                from_callback=True
            )
            return

        await query.edit_message_text(
            f"📂 گزارش دسته‌بندی\n\n"
            f"{category_name}\n\n"
            "📅 بازه زمانی را انتخاب کن:",
            reply_markup=category_report_period_keyboard()
        )

        return
    # ==========================================
    # انتخاب دسته‌بندی
    # ==========================================
    if action.startswith("cat_report_cat:"):

        category_id = int(action.split(":")[1])

        categories = get_categories()

        category_name = None

        for cat_id, cat_name in categories:
            if cat_id == category_id:
                category_name = cat_name
                break

        if not category_name:
            await query.edit_message_text(
                "❌ دسته‌بندی پیدا نشد."
            )
            return

        context.user_data.clear()

        context.user_data["category_report_category"] = category_name

        await query.edit_message_text(
            f"📂 گزارش دسته‌بندی\n\n"
            f"{category_name}\n\n"
            "📅 بازه زمانی را انتخاب کن:",
            reply_markup=category_report_period_keyboard()
        )

        return

    # ==========================================
    # بازگشت به انتخاب دسته
    # ==========================================
    if action == "cat_report_back_category":

        await category_report_show_categories(
            update,
            context,
            from_callback=True
        )

        return

    # ==========================================
    # امروز
    # ==========================================
    if action == "cat_report_today":

        today = datetime.now(TEHRAN_TZ).date()

        start_date = today.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "امروز"
        )

        return

    # ==========================================
    # این هفته
    # شنبه تا امروز
    # ==========================================
    if action == "cat_report_this_week":

        today = datetime.now(TEHRAN_TZ).date()

        days_since_saturday = (today.weekday() + 2) % 7

        start_of_week = today - timedelta(
            days=days_since_saturday
        )

        start_date = start_of_week.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "این هفته"
        )

        return

    # ==========================================
    # هفته گذشته
    # شنبه تا جمعه
    # ==========================================
    if action == "cat_report_last_week":

        today = datetime.now(TEHRAN_TZ).date()

        days_since_saturday = (today.weekday() + 2) % 7

        start_of_this_week = today - timedelta(
            days=days_since_saturday
        )

        start_of_last_week = start_of_this_week - timedelta(days=7)

        end_of_last_week = start_of_this_week - timedelta(days=1)

        start_date = start_of_last_week.strftime("%Y-%m-%d")
        end_date = end_of_last_week.strftime("%Y-%m-%d")

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "هفته گذشته"
        )

        return

    # ==========================================
    # این ماه - بر اساس ماه شمسی
    # ==========================================
    if action == "cat_report_this_month":

        today = datetime.now(TEHRAN_TZ).date()

        today_jalali = jdatetime.date.fromgregorian(
            date=today
        )

        first_day_jalali = jdatetime.date(
            today_jalali.year,
            today_jalali.month,
            1
        )

        start_date = (
            first_day_jalali
            .togregorian()
            .strftime("%Y-%m-%d")
        )

        end_date = today.strftime("%Y-%m-%d")

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "این ماه"
        )

        return

    # ==========================================
    # سه ماه اخیر
    # ==========================================
    if action == "cat_report_quarter":

        today = datetime.now(TEHRAN_TZ).date()

        start_date = (
            today - timedelta(days=90)
        ).strftime("%Y-%m-%d")

        end_date = today.strftime("%Y-%m-%d")

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "سه ماه اخیر"
        )

        return

    # ==========================================
    # بازه دلخواه
    # ==========================================
    if action == "cat_report_manual":

        context.user_data["waiting_category_report_start"] = True

        await query.edit_message_text(
            "✏️ بازه دلخواه\n\n"
            "📅 تاریخ شروع را وارد کن:\n\n"
            "شمسی:\n"
            "1405-05-01\n\n"
            "میلادی:\n"
            "2026-07-23\n\n"
            "فرمت‌های قابل قبول:\n"
            "1405/05/01\n"
            "1405.05.01"
        )

        return

    # ==========================================
    # صفحه‌بندی
    # ==========================================
    if action.startswith("cat_report_page:"):

        page = int(action.split(":")[1])

        context.user_data["category_report_page"] = page

        start_date = context.user_data.get(
            "category_report_start"
        )

        end_date = context.user_data.get(
            "category_report_end"
        )

        period_title = context.user_data.get(
            "category_report_period",
            "بازه انتخابی"
        )

        if not start_date or not end_date:
            await query.edit_message_text(
                "❌ اطلاعات گزارش پیدا نشد.\n\n"
                "لطفاً دوباره گزارش را اجرا کن."
            )
            return

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            period_title,
            page=page
        )

        return


async def category_report_show_categories(
    update,
    context,
    from_callback=False
):
    """نمایش مجدد دسته‌بندی‌ها"""

    categories = get_categories()

    buttons = []
    row = []

    for category_id, category_name in categories:

        row.append(
            InlineKeyboardButton(
                category_name,
                callback_data=f"cat_report_cat:{category_id}"
            )
        )

        if len(row) == 2:
            buttons.append(row)
            row = []

    if row:
        buttons.append(row)

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت به گزارش‌ها",
            callback_data="reports_menu"
        )
    ])

    text = (
        "📂 گزارش بر اساس دسته‌بندی\n\n"
        "دسته‌بندی موردنظر را انتخاب کن:"
    )

    if from_callback and update.callback_query:

        await update.callback_query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )

    else:

        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )

async def show_category_report(
    update,
    context,
    start_date,
    end_date,
    period_title,
    page=None
):
    """نمایش گزارش جزئیات یک دسته‌بندی"""
    user_id = update.effective_user.id
    category_name = context.user_data.get("category_report_category")
    
    if not category_name:
        text = "❌ دسته‌بندی گزارش مشخص نیست.\nلطفاً دوباره گزارش را اجرا کن."
        if update.callback_query:
            await update.callback_query.edit_message_text(text)
        else:
            await update.message.reply_text(text, reply_markup=main_keyboard())
        context.user_data.clear()
        return
        
    rows = get_category_report_expenses(user_id, category_name, start_date, end_date)
    
    # ذخیره اطلاعات برای صفحه‌بندی
    context.user_data["category_report_start"] = start_date
    context.user_data["category_report_end"] = end_date
    context.user_data["category_report_period"] = period_title
    
    if page is None:
        page = context.user_data.get("category_report_page", 0)
    context.user_data["category_report_page"] = page

    # ==========================================
    # بدون هزینه
    # ==========================================
    if not rows:
        start_jalali = to_jalali(start_date)
        end_jalali = to_jalali(end_date)
        if start_date == end_date:
            date_text = start_jalali
        else:
            date_text = f"{start_jalali} تا {end_jalali}"
            
        text = (
            f"📂 {category_name}\n"
            f"📅 {period_title}\n"
            f"📅 {date_text}\n"
            "❌ در این بازه برای این دسته هیچ هزینه‌ای ثبت نشده."
        )
        buttons = [
            [InlineKeyboardButton("🔙 انتخاب بازه", callback_data="cat_report_period_back")],
            [InlineKeyboardButton("🔙 انتخاب دسته", callback_data="cat_report_back_category")],
            [InlineKeyboardButton("🔙 بازگشت به گزارش‌ها", callback_data="reports_menu")]
        ]
        await category_report_edit_or_send(update, text, InlineKeyboardMarkup(buttons))
        return

    # ==========================================
    # صفحه‌بندی و نمایش هزینه‌ها
    # ==========================================
    limit = 5
    total_items = len(rows)
    total_pages = (total_items + limit - 1) // limit

    if page < 0:
        page = 0
    if page >= total_pages:
        page = total_pages - 1
        
    context.user_data["category_report_page"] = page
    offset = page * limit
    page_rows = rows[offset:offset + limit]
    
    total = sum(row[1] for row in rows)
    count = len(rows)
    
    start_jalali = to_jalali(start_date)
    end_jalali = to_jalali(end_date)
    
    if start_date == end_date:
        date_text = start_jalali
    else:
        date_text = f"{start_jalali} تا {end_jalali}"

    # ==========================================
    # متن گزارش
    # ==========================================
    text = (
        f"📂 {category_name}\n"
        f"📅 {period_title}\n"
        f"📆 {date_text}\n"
        f"📄 صفحه {page + 1} از {total_pages}\n"
        f"💰 مجموع: {total:,} تومان\n"
        f"🧾 تعداد: {count} مورد\n"
        "━━━━━━━━━━━━\n"
    )
    
    for display_number, (expense_id, amount, description, category, created_at) in enumerate(page_rows, start=offset + 1):
        expense_date = created_at[:10] if created_at else ""
        expense_date_jalali = to_jalali(expense_date)
        expense_time = created_at[11:16] if created_at and len(created_at) >= 16 else ""
        
        text += f"#{display_number}\n💰 {amount:,} تومان\n📝 {description}\n📅 {expense_date_jalali}"
        if expense_time:
            text += f" | 🕐 {expense_time}"
        text += "\n"
        
    text += (
        "━━━━━━━━━━━━\n"
        f"🧾 تعداد کل: {count} مورد\n"
        f"💵 مجموع کل: {total:,} تومان"
    )

    # ==========================================
    # دکمه‌های صفحه‌بندی
    # ==========================================
    buttons = []
    nav_buttons = []
    
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("⬅️ قبلی", callback_data=f"cat_report_page:{page - 1}"))
    
    nav_buttons.append(InlineKeyboardButton(f"📄 {page + 1}/{total_pages}", callback_data="ignore"))
    
    if page < total_pages - 1:
        nav_buttons.append(InlineKeyboardButton("➡️ بعدی", callback_data=f"cat_report_page:{page + 1}"))
        
    buttons.append(nav_buttons)
    buttons.append([InlineKeyboardButton("🔙 انتخاب بازه", callback_data="cat_report_period_back")])
    buttons.append([InlineKeyboardButton("🔙 انتخاب دسته", callback_data="cat_report_back_category")])
    buttons.append([InlineKeyboardButton("🔙 بازگشت به گزارش‌ها", callback_data="reports_menu")])
    
    await category_report_edit_or_send(update, text, InlineKeyboardMarkup(buttons))



async def category_report_edit_or_send(
    update,
    text,
    reply_markup
):
    """ارسال یا ویرایش پیام گزارش دسته‌بندی"""

    if update.callback_query:

        await update.callback_query.edit_message_text(
            text,
            reply_markup=reply_markup
        )

    else:

        await update.message.reply_text(
            text,
            reply_markup=reply_markup
        )
async def render_edit_delete_menu(update, context, page):
    """نمایش منوی حذف/ویرایش هزینه‌ها (ارسال پیام جدید یا ویرایش پیام فعلی)"""
    user_id = update.effective_user.id
    limit = 5  # تعداد آیتم در هر صفحه

    offset = page * limit
    response = (
        supabase.table("expenses")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .range(offset, offset + limit - 1)
        .execute()
    )

    rows = [(row["id"], row["amount"], row["description"], row["category"], row["created_at"]) for row in response.data]

    # بررسی وجود صفحه بعدی
    next_check = (
        supabase.table("expenses")
        .select("id")
        .eq("user_id", user_id)
        .range(offset + limit, offset + limit)
        .execute()
    )
    has_next = len(next_check.data) > 0

    if not rows and page == 0:
        if update.callback_query:
            await update.callback_query.edit_message_text("📋 هنوز هزینه‌ای ثبت نشده.")
        else:
            await update.message.reply_text("📋 هنوز هزینه‌ای ثبت نشده.", reply_markup=main_keyboard())
        return

    if not rows:
        if update.callback_query:
            await update.callback_query.edit_message_text(
                "📋 صفحه خالی است.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت", callback_data="back_menu")]])
            )
        else:
            await update.message.reply_text("📋 صفحه خالی است.", reply_markup=main_keyboard())
        return

    # ساخت دکمه‌ها
    buttons = []
    for display_number, (expense_id, amount, description, category, created_at) in enumerate(rows, start=offset + 1):
        buttons.append([
            InlineKeyboardButton(
                f"✏️ #{display_number} | {amount:,} تومان",
                callback_data=f"edit:{expense_id}"
            ),
            InlineKeyboardButton(
                "🗑️ حذف",
                callback_data=f"delete:{expense_id}"
            ),
        ])

    # دکمه‌های صفحه‌بندی
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("⬅️ قبلی", callback_data=f"edit_page:{page-1}"))

    nav_buttons.append(InlineKeyboardButton(f"📄 {page + 1}", callback_data="ignore"))

    if has_next:
        nav_buttons.append(InlineKeyboardButton("➡️ بعدی", callback_data=f"edit_page:{page+1}"))

    buttons.append(nav_buttons)
    buttons.append([InlineKeyboardButton("🔙 بازگشت به منو", callback_data="back_menu")])

    # ذخیره صفحه فعلی در context
    context.user_data["edit_page"] = page

    text = (
        f"🗑️ حذف / ✏️ ویرایش\n"
        f"📄 صفحه {page + 1}\n"
        f"📋 {len(rows)} هزینه در این صفحه\n\n"
        "هزینه موردنظر را انتخاب کن:"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(buttons))


async def edit_delete_menu(update, context):
    """نمایش هزینه‌ها با صفحه‌بندی برای حذف/ویرایش"""
    user_id = update.effective_user.id
    if not is_allowed(user_id):
        return

    page = context.user_data.get("edit_page", 0)
    await render_edit_delete_menu(update, context, page)


async def delete_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    expense_id = int(query.data.split(":")[1])
    expense = get_expense(user_id, expense_id)
    if not expense:
        await query.edit_message_text("❌ هزینه پیدا نشد.")
        return
    _, amount, description, category, _ = expense
    buttons = [
        [
            InlineKeyboardButton("✅ بله، حذف کن", callback_data=f"confirm_delete:{expense_id}"),
            InlineKeyboardButton("❌ انصراف", callback_data="cancel_delete"),
        ]
    ]
    await query.edit_message_text(
        f"⚠️ حذف این هزینه؟\n\n{category}\n💰 {amount:,} تومان\n📝 {description}",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def confirm_delete_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    expense_id = int(query.data.split(":")[1])
    deleted = delete_expense(user_id, expense_id)
    if deleted:
        await query.edit_message_text(f"✅ هزینه #{expense_id} حذف شد.")
    else:
        await query.edit_message_text("❌ هزینه پیدا نشد.")

async def cancel_delete_callback(update, context):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("❌ حذف لغو شد.")

async def edit_callback(update, context):
    """ویرایش هزینه انتخاب شده"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    
    expense_id = int(query.data.split(":")[1])
    expense = get_expense(user_id, expense_id)
    if not expense:
        await query.edit_message_text("❌ هزینه پیدا نشد.")
        return
    
    _, amount, description, category, _ = expense
    context.user_data.clear()
    context.user_data["editing_expense"] = expense_id
    context.user_data["editing_category"] = category
    context.user_data["waiting_for_edit"] = True
    
    await query.edit_message_text(
        f"✏️ ویرایش هزینه #{expense_id}\n\n"
        f"{category}\n"
        f"💰 مبلغ فعلی: {amount:,} تومان\n"
        f"📝 {description}\n\n"
        "مبلغ و توضیح جدید را بفرست.\n\n"
        "مثال:\n95000 ناهار رستوران"
    )
    await context.bot.send_message(
        chat_id=user_id,
        text="🔙 بازگشت",
        reply_markup=back_keyboard()
    )

async def edit_page_callback(update, context):
    """تغییر صفحه در منوی حذف/ویرایش"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    page = int(query.data.split(":")[1])
    context.user_data["edit_page"] = page
    await render_edit_delete_menu(update, context, page)


def settings_markup():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏷️ دسته‌بندی‌ها", callback_data="manage_categories")],
        [InlineKeyboardButton("🔑 کلمات دسته‌بندی", callback_data="manage_keywords")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="back_menu")],
    ])


async def settings(update, context):
    await update.message.reply_text(
        "⚙️ تنظیمات",
        reply_markup=settings_markup()
    )

async def manage_categories(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    context.user_data.clear()
    categories = get_categories()
    text = "🏷️ مدیریت دسته‌بندی‌ها\n\n"
    for category_id, name in categories:
        text += f"• {name}\n"
    buttons = [
        [InlineKeyboardButton("➕ افزودن دسته", callback_data="category_add")],
        [InlineKeyboardButton("✏️ تغییر نام", callback_data="category_rename"), InlineKeyboardButton("🗑️ حذف", callback_data="category_delete")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="settings_menu")],
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))
def keywords_menu_markup():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ افزودن کلمه", callback_data="keyword_add")],
        [InlineKeyboardButton("✏️ ویرایش کلمه", callback_data="keyword_edit")],
        [InlineKeyboardButton("🗑️ حذف کلمه", callback_data="keyword_delete")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="settings_menu")],
    ])


async def build_keywords_menu_text():
    """متن منوی مدیریت کلمات کلیدی با لیست کلمات هر دسته"""
    categories = get_categories()
    text = "🔑 مدیریت کلمات دسته‌بندی\n\n"

    for category_id, name in categories:
        response = (
            supabase
            .table("category_keywords")
            .select("keyword")
            .eq("category_id", category_id)
            .execute()
        )

        keywords = [
            row["keyword"]
            for row in response.data
            if row.get("keyword")
        ]

        text += f"{name}\n"

        if keywords:
            text += "  " + "، ".join(keywords) + "\n"
        else:
            text += "  — کلمه‌ای ثبت نشده\n"

        text += "\n"

    return text


async def manage_keywords(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    text = await build_keywords_menu_text()

    await query.edit_message_text(
        text,
        reply_markup=keywords_menu_markup()
    )
async def keyword_add_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    categories = get_categories()

    buttons = []
    for category_id, name in categories:
        buttons.append([
            InlineKeyboardButton(
                name,
                callback_data=f"keyword_add_cat:{category_id}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت",
            callback_data="manage_keywords"
        )
    ])

    await query.edit_message_text(
        "➕ افزودن کلمه\n\n"
        "کلمه را می‌خواهی به کدام دسته اضافه کنی؟",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_edit_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    categories = get_categories()

    if not categories:
        await query.edit_message_text(
            "✏️ هیچ دسته‌بندی‌ای وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت",
                    callback_data="manage_keywords"
                )]
            ])
        )
        return

    buttons = []

    for category_id, name in categories:
        buttons.append([
            InlineKeyboardButton(
                name,
                callback_data=f"keyword_edit_category:{category_id}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت",
            callback_data="manage_keywords"
        )
    ])

    await query.edit_message_text(
        "✏️ دسته‌بندی موردنظر برای ویرایش را انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_edit_category_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    category_id = int(query.data.split(":")[1])

    response = (
        supabase
        .table("category_keywords")
        .select("id, keyword")
        .eq("category_id", category_id)
        .order("id")
        .execute()
    )

    keywords = response.data or []

    if not keywords:
        await query.edit_message_text(
            "✏️ برای این دسته کلمه‌ای وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت به دسته‌بندی‌ها",
                    callback_data="keyword_edit"
                )]
            ])
        )
        return

    buttons = []

    for row in keywords:
        buttons.append([
            InlineKeyboardButton(
                row["keyword"],
                callback_data=f"keyword_edit_select:{row['id']}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت به دسته‌بندی‌ها",
            callback_data="keyword_edit"
        )
    ])

    await query.edit_message_text(
        "✏️ کلمه‌ای که می‌خواهی ویرایش کنی را انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_edit_select_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    keyword_id = int(query.data.split(":")[1])

    context.user_data["keyword_edit_id"] = keyword_id

    await query.edit_message_text(
        "✏️ ویرایش کلمه\n\n"
        "نام جدید کلمه را وارد کن:\n\n"
        "مثال: برگر"
    )
async def keyword_delete_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    categories = get_categories()

    if not categories:
        await query.edit_message_text(
            "🗑️ هیچ دسته‌بندی‌ای وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت",
                    callback_data="manage_keywords"
                )]
            ])
        )
        return

    buttons = []

    for category_id, name in categories:
        buttons.append([
            InlineKeyboardButton(
                name,
                callback_data=f"keyword_delete_category:{category_id}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت",
            callback_data="manage_keywords"
        )
    ])

    await query.edit_message_text(
        "🗑️ دسته‌بندی موردنظر برای حذف را انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_delete_category_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    category_id = int(query.data.split(":")[1])

    response = (
        supabase
        .table("category_keywords")
        .select("id, keyword")
        .eq("category_id", category_id)
        .order("id")
        .execute()
    )

    keywords = response.data or []

    if not keywords:
        await query.edit_message_text(
            "🗑️ برای این دسته کلمه‌ای وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت به دسته‌بندی‌ها",
                    callback_data="keyword_delete"
                )]
            ])
        )
        return

    buttons = []

    for row in keywords:
        buttons.append([
            InlineKeyboardButton(
                f"🗑️ {row['keyword']}",
                callback_data=f"keyword_delete_select:{row['id']}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت به دسته‌بندی‌ها",
            callback_data="keyword_delete"
        )
    ])

    await query.edit_message_text(
        "🗑️ کلمه‌ای که می‌خواهی حذف کنی را انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_delete_select_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    keyword_id = int(query.data.split(":")[1])

    response = (
        supabase
        .table("category_keywords")
        .select("keyword")
        .eq("id", keyword_id)
        .execute()
    )

    if not response.data:
        await query.edit_message_text(
            "❌ این کلمه پیدا نشد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="manage_keywords")]
            ])
        )
        return

    keyword = response.data[0]["keyword"]

    buttons = [
        [
            InlineKeyboardButton(
                "✅ بله، حذف کن",
                callback_data=f"keyword_delete_confirm:{keyword_id}"
            ),
            InlineKeyboardButton(
                "❌ لغو",
                callback_data="manage_keywords"
            )
        ]
    ]

    await query.edit_message_text(
        f"⚠️ مطمئنی می‌خواهی «{keyword}» حذف شود؟",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
async def keyword_delete_confirm_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    keyword_id = int(query.data.split(":")[1])

    response = (
        supabase
        .table("category_keywords")
        .delete()
        .eq("id", keyword_id)
        .execute()
    )

    if response.data:
        await query.edit_message_text(
            "✅ کلمه با موفقیت حذف شد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت به کلمات",
                    callback_data="manage_keywords"
                )]
            ])
        )
    else:
        await query.edit_message_text(
            "❌ حذف کلمه انجام نشد.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    "🔙 بازگشت",
                    callback_data="manage_keywords"
                )]
            ])
        )
async def keyword_add_category_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    category_id = int(query.data.split(":")[1])

    context.user_data["keyword_add_category_id"] = category_id

    await query.edit_message_text(
        "➕ افزودن کلمه\n\n"
        "کلمه جدید را وارد کن:\n\n"
        "مثال: پیتزا"
    )    
async def ignore_callback(update, context):
    """دکمه‌های غیرفعال (شماره صفحه)"""
    query = update.callback_query
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    await query.answer("📄 این دکمه فقط نمایشی است")

async def category_add_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    context.user_data.clear()
    context.user_data["waiting_category_add"] = True
    await query.edit_message_text("➕ افزودن دسته\n\nنام دسته جدید را بفرست.\n\nمثال:\n☕ کافه\n\n🔙 برای لغو، بازگشت را بزن.")
    await context.bot.send_message(chat_id=query.from_user.id, text="🔙 بازگشت", reply_markup=back_keyboard())

async def category_rename_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    categories = get_categories()
    buttons = []
    for category_id, name in categories:
        buttons.append([InlineKeyboardButton(name, callback_data=f"rename_select:{category_id}")])
    buttons.append([InlineKeyboardButton("🔙 بازگشت", callback_data="manage_categories")])
    await query.edit_message_text("✏️ کدام دسته را می‌خواهی تغییر نام بدهی؟", reply_markup=InlineKeyboardMarkup(buttons))

async def rename_select_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    category_id = int(query.data.split(":")[1])
    context.user_data.clear()
    context.user_data["waiting_category_rename"] = True
    context.user_data["rename_category_id"] = category_id
    await query.edit_message_text("✏️ نام جدید دسته را بفرست.\n\nمثال:\n☕ کافه")
    await context.bot.send_message(chat_id=query.from_user.id, text="🔙 بازگشت", reply_markup=back_keyboard())

async def category_delete_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    categories = get_categories()
    buttons = []
    for category_id, name in categories:
        if name == "📦 سایر":
            continue
        buttons.append([InlineKeyboardButton(f"🗑️ {name}", callback_data=f"delete_category:{category_id}")])
    buttons.append([InlineKeyboardButton("🔙 بازگشت", callback_data="manage_categories")])
    await query.edit_message_text(
        "🗑️ کدام دسته را می‌خواهی حذف کنی؟\n\n⚠️ هزینه‌های آن دسته به «📦 سایر» منتقل می‌شوند.",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def delete_category_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    category_id = int(query.data.split(":")[1])
    categories = get_categories()
    category_name = None
    for cid, name in categories:
        if cid == category_id:
            category_name = name
            break
    if not category_name:
        await query.edit_message_text("❌ دسته پیدا نشد.")
        return
    buttons = [
        [
            InlineKeyboardButton("✅ بله، حذف کن", callback_data=f"confirm_category_delete:{category_id}"),
            InlineKeyboardButton("❌ انصراف", callback_data="manage_categories"),
        ]
    ]
    await query.edit_message_text(
        f"⚠️ حذف دسته «{category_name}»؟\n\nهزینه‌های قبلی این دسته به «📦 سایر» منتقل می‌شوند.",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def confirm_category_delete_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return
    category_id = int(query.data.split(":")[1])
    deleted = delete_category(category_id)
    if deleted:
        await query.edit_message_text("✅ دسته حذف شد.\nهزینه‌های قبلی آن به «📦 سایر» منتقل شدند.")
    else:
        await query.edit_message_text("❌ این دسته قابل حذف نیست.")

async def back_callback(update, context):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    context.user_data.clear()

    await query.edit_message_text(
        "🏠 منوی اصلی"
    )

    await context.bot.send_message(
        chat_id=query.from_user.id,
        text="از منوی پایین انتخاب کن 👇",
        reply_markup=main_keyboard()
    )

async def settings_menu_callback(update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    await query.edit_message_text(
        "⚙️ تنظیمات",
        reply_markup=settings_markup()
    )

# ==========================================
# توابع مدیریت هزینه‌های سریع
# ==========================================
async def quick_add_callback(update, context):
    """شروع افزودن هزینه سریع جدید"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    # پاک کردن تمام stateهای قبلی
    context.user_data.clear()

    categories = get_categories()

    category_buttons = []
    row = []

    for cat_id, cat_name in categories:
        row.append(
            InlineKeyboardButton(
                cat_name,
                callback_data=f"quick_add_cat_{cat_id}"
            )
        )

        if len(row) == 2:
            category_buttons.append(row)
            row = []

    if row:
        category_buttons.append(row)

    category_buttons.append([
        InlineKeyboardButton(
            "🔙 بازگشت",
            callback_data="quick_manage"
        )
    ])

    await query.edit_message_text(
        "➕ افزودن هزینه سریع جدید\n\n"
        "۱. دسته‌بندی را انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(category_buttons)
    )

async def quick_add_category_callback(update, context):
    """انتخاب دسته‌بندی برای هزینه سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    try:
        cat_id = int(query.data.replace("quick_add_cat_", ""))
    except ValueError:
        await query.edit_message_text("❌ دسته‌بندی نامعتبر است.")
        return

    categories = get_categories()

    category_name = None

    for cid, name in categories:
        if cid == cat_id:
            category_name = name
            break

    if not category_name:
        await query.edit_message_text("❌ دسته‌بندی پیدا نشد.")
        return

    # فقط state مربوط به هزینه سریع را نگه می‌داریم
    context.user_data.clear()
    context.user_data["quick_add_category"] = category_name
    context.user_data["waiting_quick_add"] = True

    await query.edit_message_text(
        f"➕ افزودن هزینه سریع\n\n"
        f"📂 دسته‌بندی: {category_name}\n\n"
        "۲. مبلغ را وارد کن.\n\n"
        "مثال:\n"
        "85000\n\n"
        "یا همراه با نام:\n"
        "85000 ناهار",
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton(
                    "🔙 بازگشت",
                    callback_data="quick_manage"
                )
            ]
        ])
    )
async def quick_delete_ask_callback(update, context):
    """نمایش تأیید حذف هزینه سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    try:
        quick_id = int(
            query.data.replace("quick_delete_ask_", "")
        )
    except ValueError:
        await query.edit_message_text(
            "❌ شناسه هزینه نامعتبر است."
        )
        return

    response = (
        supabase
        .table("quick_expenses")
        .select("*")
        .eq("id", quick_id)
        .eq("user_id", user_id)
        .execute()
    )

    if not response.data:
        await query.edit_message_text(
            "❌ هزینه سریع پیدا نشد.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "🔙 بازگشت",
                        callback_data="quick_manage"
                    )
                ]
            ])
        )
        return

    item = response.data[0]

    keyboard = [
        [
            InlineKeyboardButton(
                "✅ بله، حذف کن",
                callback_data=f"quick_delete_confirm_{quick_id}"
            ),
            InlineKeyboardButton(
                "❌ انصراف",
                callback_data="quick_manage"
            )
        ]
    ]

    await query.edit_message_text(
        "⚠️ حذف هزینه سریع\n\n"
        f"📝 {item['name']}\n"
        f"💰 {item['amount']:,} تومان\n"
        f"📂 {item['category']}\n\n"
        "آیا مطمئنی که می‌خواهی این هزینه را حذف کنی؟",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
async def quick_delete_confirm_callback(update, context):
    """تأیید حذف هزینه سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    # دریافت id هزینه سریع
    try:
        quick_id = int(
            query.data.replace("quick_delete_confirm_", "")
        )
    except ValueError:
        await query.edit_message_text(
            "❌ شناسه هزینه نامعتبر است.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "🔙 بازگشت",
                        callback_data="quick_manage"
                    )
                ]
            ])
        )
        return

    # حذف از دیتابیس
    deleted = delete_quick_expense(user_id, quick_id)

    if deleted:
        context.user_data.clear()

        await query.edit_message_text(
            "✅ هزینه سریع حذف شد.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "⚙️ مدیریت هزینه‌های سریع",
                        callback_data="quick_manage"
                    )
                ],
                [
                    InlineKeyboardButton(
                        "🧾 هزینه‌های سریع",
                        callback_data="quick_menu"
                    )
                ],
                [
                    InlineKeyboardButton(
                        "🔙 منوی اصلی",
                        callback_data="back_menu"
                    )
                ]
            ])
        )

    else:
        await query.edit_message_text(
            "❌ خطا در حذف هزینه سریع.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "🔙 بازگشت",
                        callback_data="quick_manage"
                    )
                ]
            ])
        )
async def quick_edit_select_callback(update, context):
    """انتخاب هزینه برای ویرایش"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    try:
        quick_id = int(
            query.data.replace("quick_edit_select_", "")
        )
    except ValueError:
        await query.edit_message_text(
            "❌ شناسه هزینه نامعتبر است.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "🔙 بازگشت",
                        callback_data="quick_manage"
                    )
                ]
            ])
        )
        return

    response = (
        supabase
        .table("quick_expenses")
        .select("*")
        .eq("id", quick_id)
        .eq("user_id", user_id)
        .execute()
    )

    if not response.data:
        await query.edit_message_text(
            "❌ هزینه سریع پیدا نشد.",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(
                        "🔙 بازگشت",
                        callback_data="quick_manage"
                    )
                ]
            ])
        )
        return

    item = response.data[0]

    # پاک کردن state قبلی
    context.user_data.clear()

    context.user_data["quick_edit_id"] = quick_id
    context.user_data["quick_edit_name"] = item["name"]
    context.user_data["quick_edit_category"] = item["category"]
    context.user_data["waiting_quick_edit"] = True

    await query.edit_message_text(
        f"✏️ ویرایش «{item['name']}»\n\n"
        "مبلغ جدید را وارد کن:\n\n"
        f"💰 مبلغ فعلی: {item['amount']:,} تومان\n"
        f"📂 دسته‌بندی: {item['category']}\n\n"
        "مثال:\n"
        "75000\n\n"
        "🔹 فقط عدد وارد کن (نام و دسته‌بندی تغییر نمی‌کنند).",
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton(
                    "🔙 بازگشت",
                    callback_data="quick_manage"
                )
            ]
        ])
    )

async def quick_edit_callback(update, context):
    """ویرایش هزینه سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    # دریافت هزینه‌های سریع از دیتابیس
    quick_items = get_quick_expenses(user_id)

    if not quick_items:
        await query.edit_message_text(
            "📋 **هیچ هزینه سریعی برای ویرایش وجود ندارد.**\n\n"
            "ابتدا از طریق «➕ افزودن هزینه سریع» یک هزینه اضافه کن.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="quick_manage")]
            ])
        )
        return

    buttons = []
    for item in quick_items:
        buttons.append([
            InlineKeyboardButton(
                f"✏️ {item['name']} ({item['amount']:,})",
                callback_data=f"quick_edit_select_{item['id']}"
            )
        ])

    buttons.append([InlineKeyboardButton("🔙 بازگشت", callback_data="quick_manage")])

    await query.edit_message_text(
        "✏️ **ویرایش هزینه سریع**\n\n"
        "هزینه‌ای که میخوای ویرایش کنی رو انتخاب کن:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def quick_delete_callback(update, context):
    """حذف هزینه سریع"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not is_allowed(user_id):
        return

    # دریافت هزینه‌های سریع از دیتابیس
    quick_items = get_quick_expenses(user_id)

    if not quick_items:
        await query.edit_message_text(
            "📋 **هیچ هزینه سریعی برای حذف وجود ندارد.**",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="quick_manage")]
            ])
        )
        return

    buttons = []
    for item in quick_items:
        buttons.append([
            InlineKeyboardButton(
                f"🗑️ {item['name']} ({item['amount']:,})",
                callback_data=f"quick_delete_ask_{item['id']}"
            )
        ])

    buttons.append([InlineKeyboardButton("🔙 بازگشت", callback_data="quick_manage")])

    await query.edit_message_text(
        "🗑️ **حذف هزینه سریع**\n\n"
        "هزینه‌ای که میخوای حذف کنی رو انتخاب کن:\n\n"
        "⚠️ فقط از لیست هزینه‌های سریع حذف میشه.",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

# ==========================================
# تابع خروجی اکسل
# ==========================================
async def export_excel_callback(update, context):
    """خروجی اکسل از طریق دکمه گزارش‌ها"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_allowed(user_id):
        return

    # تابع اصلی خروجی اکسل را اجرا می‌کنیم
    await export_excel(update, context, from_callback=True)


async def export_excel(update, context, from_callback=False):
    user_id = update.effective_user.id

    if not is_allowed(user_id):
        await update.message.reply_text(
            "⛔ شما اجازه استفاده از این بخش را ندارید."
        )
        return

    current_month = datetime.now(TEHRAN_TZ)
    month = current_month.strftime("%Y-%m")

    try:
        # ==========================================
        # محاسبه بازه ماه
        # ==========================================
        start_date = f"{month}-01 00:00:00"

        if current_month.month == 12:
            next_month = current_month.replace(
                year=current_month.year + 1,
                month=1,
                day=1
            )
        else:
            next_month = current_month.replace(
                month=current_month.month + 1,
                day=1
            )

        end_date = next_month.strftime("%Y-%m-%d 00:00:00")

        logger.info(
            f"Export Excel | user={user_id} | "
            f"start={start_date} | end={end_date}"
        )

        # ==========================================
        # دریافت هزینه‌ها از Supabase
        # ==========================================
        response = (
            supabase
            .table("expenses")
            .select("*")
            .eq("user_id", user_id)
            .gte("created_at", start_date)
            .lt("created_at", end_date)
            .order("created_at", desc=True)
            .execute()
        )

        rows = response.data or []

        if not rows:
            await update.message.reply_text(
                "📊 این ماه هنوز هزینه‌ای ثبت نشده.",
                reply_markup=main_keyboard()
            )
            return

        # ==========================================
        # Importهای مربوط به Excel
        # ==========================================
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # ==========================================
        # ساخت Workbook
        # ==========================================
        wb = Workbook()

        # ==========================================
        # شیت اول: هزینه‌ها
        # ==========================================
        ws = wb.active
        ws.title = "هزینه‌ها"

        # ==========================================
        # عنوان فایل
        # ==========================================
                # تبدیل ماه به شمسی برای عنوان
        month_parts = month.split('-')
        year, month_num = int(month_parts[0]), int(month_parts[1])
        gregorian_date = datetime(year, month_num, 1)
        jalali_date = jdatetime.date.fromgregorian(date=gregorian_date)
        jalali_month = f"{jalali_date.year:04d}-{jalali_date.month:02d}"
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"💰 گزارش هزینه‌های ماه {jalali_month}"

        ws["A1"].font = Font(
            bold=True,
            size=18
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        ws.row_dimensions[1].height = 35

        # ==========================================
        # اطلاعات خلاصه
        # ==========================================
        total = sum(int(row.get("amount", 0)) for row in rows)
        count = len(rows)

        ws["A2"] = "تعداد هزینه‌ها"
        ws["B2"] = count

        ws["D2"] = "مجموع هزینه"
        ws["E2"] = total

        ws["A2"].font = Font(bold=True)
        ws["D2"].font = Font(bold=True)

        ws["B2"].font = Font(bold=True)
        ws["E2"].font = Font(bold=True)

        # فرمت مبلغ
        ws["E2"].number_format = '#,##0" تومان"'

        # ==========================================
        # هدر جدول
        # ==========================================
        headers = [
            "ردیف",
            "دسته‌بندی",
            "مبلغ (تومان)",
            "توضیحات",
            "تاریخ",
            "ساعت"
        ]

        header_row = 4

        for col, header in enumerate(headers, 1):
            cell = ws.cell(
                row=header_row,
                column=col,
                value=header
            )

            cell.font = Font(
                bold=True,
                size=11
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ==========================================
        # وارد کردن اطلاعات
        # ==========================================
        for i, row in enumerate(rows, start=1):

            created_at = str(row.get("created_at", ""))

            # استفاده از تابع تبدیل به شمسی
            date_part = to_jalali(created_at)

            time_part = (
                created_at[11:16]
                if len(created_at) >= 16
                else ""
            )

            amount = int(row.get("amount", 0))

            excel_row = header_row + i

            ws.cell(excel_row, 1, i)
            ws.cell(excel_row, 2, row.get("category", ""))
            ws.cell(excel_row, 3, amount)
            ws.cell(excel_row, 4, row.get("description", ""))
            ws.cell(excel_row, 5, date_part)
            ws.cell(excel_row, 6, time_part)

            # فرمت مبلغ
            ws.cell(
                excel_row,
                3
            ).number_format = '#,##0" تومان"'

            # تراز وسط برای اطلاعات عددی
            ws.cell(excel_row, 1).alignment = Alignment(
                horizontal="center"
            )

            ws.cell(excel_row, 3).alignment = Alignment(
                horizontal="center"
            )

            ws.cell(excel_row, 5).alignment = Alignment(
                horizontal="center"
            )

            ws.cell(excel_row, 6).alignment = Alignment(
                horizontal="center"
            )

            # راست‌چین برای متن
            ws.cell(excel_row, 2).alignment = Alignment(
                horizontal="right"
            )

            ws.cell(excel_row, 4).alignment = Alignment(
                horizontal="right"
            )

        # ==========================================
        # ردیف جمع کل
        # ==========================================
        total_row = header_row + len(rows) + 2

        ws.cell(total_row, 2, "💰 جمع کل")
        ws.cell(total_row, 3, total)

        ws.cell(total_row, 2).font = Font(
            bold=True,
            size=12
        )

        ws.cell(total_row, 3).font = Font(
            bold=True,
            size=12
        )

        ws.cell(
            total_row,
            3
        ).number_format = '#,##0" تومان"'

        # ==========================================
        # جدول Excel
        # ==========================================
        table_end_row = header_row + len(rows)

        table_ref = f"A{header_row}:F{table_end_row}"

        tab = Table(
            displayName="ExpensesTable",
            ref=table_ref
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tab.tableStyleInfo = style
        ws.add_table(tab)

        # ==========================================
        # فریز کردن هدر
        # ==========================================
        ws.freeze_panes = "A5"

        # ==========================================
        # فیلتر
        # ==========================================
        ws.auto_filter.ref = table_ref

        # ==========================================
        # عرض ستون‌ها
        # ==========================================
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 12

        # ==========================================
        # راست‌چین و Wrap Text
        # ==========================================
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "right",
                    vertical="center",
                    wrap_text=True
                )

        # ==========================================
        # شیت دوم: گزارش ماه
        # ==========================================
        report_ws = wb.create_sheet("گزارش ماه")

        # استفاده از jalali_month که قبلاً محاسبه شد
        report_ws.merge_cells("A1:D1")
        report_ws["A1"] = f"📊 خلاصه هزینه‌های ماه {jalali_month}"

        report_ws["A1"].font = Font(
            bold=True,
            size=18
        )

        report_ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        report_ws.row_dimensions[1].height = 35

        # ==========================================
        # اطلاعات کلی
        # ==========================================
        report_ws["A3"] = "مجموع هزینه‌ها"
        report_ws["B3"] = total

        report_ws["A4"] = "تعداد هزینه‌ها"
        report_ws["B4"] = count

        average = total // count if count else 0

        report_ws["A5"] = "میانگین هر هزینه"
        report_ws["B5"] = average

        for cell in ["A3", "A4", "A5"]:
            report_ws[cell].font = Font(bold=True)

        for cell in ["B3", "B5"]:
            report_ws[cell].number_format = '#,##0" تومان"'

        # ==========================================
        # محاسبه دسته‌بندی‌ها
        # ==========================================
        categories = {}

        for row in rows:

            category = row.get(
                "category",
                "📦 سایر"
            )

            amount = int(
                row.get("amount", 0)
            )

            if category not in categories:
                categories[category] = {
                    "total": 0,
                    "count": 0
                }

            categories[category]["total"] += amount
            categories[category]["count"] += 1

        # مرتب‌سازی از بیشترین هزینه
        sorted_categories = sorted(
            categories.items(),
            key=lambda x: x[1]["total"],
            reverse=True
        )

        # ==========================================
        # جدول دسته‌بندی‌ها
        # ==========================================
        category_header_row = 8

        category_headers = [
            "دسته‌بندی",
            "مجموع (تومان)",
            "تعداد",
            "درصد از کل"
        ]

        for col, header in enumerate(
            category_headers,
            start=1
        ):

            cell = report_ws.cell(
                category_header_row,
                col,
                header
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # ==========================================
        # اطلاعات دسته‌ها
        # ==========================================
        for i, (category, data) in enumerate(
            sorted_categories,
            start=1
        ):

            row_num = category_header_row + i

            category_total = data["total"]
            category_count = data["count"]

            percentage = (
                category_total / total
                if total > 0
                else 0
            )

            report_ws.cell(
                row_num,
                1,
                category
            )

            report_ws.cell(
                row_num,
                2,
                category_total
            )

            report_ws.cell(
                row_num,
                3,
                category_count
            )

            report_ws.cell(
                row_num,
                4,
                percentage
            )

            report_ws.cell(
                row_num,
                2
            ).number_format = '#,##0" تومان"'

            report_ws.cell(
                row_num,
                4
            ).number_format = "0.00%"

        # ==========================================
        # جدول دسته‌بندی
        # ==========================================
        category_end_row = (
            category_header_row +
            len(sorted_categories)
        )

        category_table = Table(
            displayName="CategoryTable",
            ref=f"A{category_header_row}:D{category_end_row}"
        )

        category_style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        category_table.tableStyleInfo = category_style

        report_ws.add_table(category_table)

        # ==========================================
        # نمودار دایره‌ای
        # ==========================================
        if len(sorted_categories) > 0:

            pie = PieChart()

            labels = Reference(
                report_ws,
                min_col=1,
                min_row=category_header_row + 1,
                max_row=category_end_row
            )

            data = Reference(
                report_ws,
                min_col=2,
                min_row=category_header_row,
                max_row=category_end_row
            )

            pie.add_data(
                data,
                titles_from_data=True
            )

            pie.set_categories(labels)

            pie.title = "سهم هزینه‌ها بر اساس دسته‌بندی"

            pie.height = 8
            pie.width = 12

            report_ws.add_chart(
                pie,
                "F3"
            )

        # ==========================================
        # نمودار میله‌ای
        # ==========================================
        if len(sorted_categories) > 0:

            bar = BarChart()

            data = Reference(
                report_ws,
                min_col=2,
                min_row=category_header_row,
                max_row=category_end_row
            )

            labels = Reference(
                report_ws,
                min_col=1,
                min_row=category_header_row + 1,
                max_row=category_end_row
            )

            bar.add_data(
                data,
                titles_from_data=True
            )

            bar.set_categories(labels)

            bar.title = "مقایسه هزینه دسته‌بندی‌ها"
            bar.y_axis.title = "مبلغ"
            bar.x_axis.title = "دسته‌بندی"

            bar.height = 8
            bar.width = 14

            report_ws.add_chart(
                bar,
                "F20"
            )

        # ==========================================
        # عرض ستون‌های گزارش
        # ==========================================
        report_ws.column_dimensions["A"].width = 25
        report_ws.column_dimensions["B"].width = 22
        report_ws.column_dimensions["C"].width = 12
        report_ws.column_dimensions["D"].width = 18

        # ==========================================
        # Freeze
        # ==========================================
        report_ws.freeze_panes = "A9"

        # ==========================================
        # ذخیره فایل در حافظه
        # ==========================================
        output = BytesIO()

        wb.save(output)

        output.seek(0)

        filename = f"گزارش_هزینه_{jalali_month}.xlsx"

        # ==========================================
        # ارسال فایل
        # ==========================================
        caption = (
            f"📊 گزارش کامل هزینه‌های ماه {jalali_month}\n\n"
            f"🧾 تعداد: {count} مورد\n"
            f"💰 مجموع: {total:,} تومان\n"
            f"📊 میانگین: {average:,} تومان"
        )

        if from_callback:
            await update.callback_query.message.reply_document(
                document=output,
                filename=filename,
                caption=caption,
                reply_markup=main_keyboard()
            )
        else:
            await update.message.reply_document(
                document=output,
                filename=filename,
                caption=caption,
                reply_markup=main_keyboard()
            )

        logger.info(
            f"Export Excel successful | "
            f"user={user_id} | "
            f"rows={count} | "
            f"total={total}"
        )

    except Exception as e:
        logger.exception(
            f"خطا در خروجی اکسل برای user={user_id}"
        )

        error_text = f"❌ خطا در ایجاد فایل اکسل:\n\n{str(e)}"

        if from_callback:
            await update.callback_query.message.reply_text(
                error_text,
                reply_markup=main_keyboard()
            )
        else:
            await update.message.reply_text(
                error_text,
                reply_markup=main_keyboard()
            )
# ==========================================
# هندلر اصلی پیام‌ها
# ==========================================
REPORTS_MENU_TEXT = "📊 گزارش‌ها\n\nنوع گزارش موردنظر را انتخاب کن:"


def reports_menu_markup():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 خلاصه مالی", callback_data="report_stats")],
        [InlineKeyboardButton("📅 گزارش بر اساس تاریخ", callback_data="report_advanced")],
        [InlineKeyboardButton("📂 گزارش بر اساس دسته‌بندی", callback_data="report_category")],
        [InlineKeyboardButton("📋 لیست هزینه‌ها", callback_data="report_recent")],
        [InlineKeyboardButton("📥 خروجی اکسل", callback_data="report_excel")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="back_menu")],
    ])


async def reports_menu(update, context):
    """منوی گزارش‌ها"""
    user_id = update.effective_user.id

    if not is_allowed(user_id):
        return

    await update.message.reply_text(
        REPORTS_MENU_TEXT,
        reply_markup=reports_menu_markup()
    )

async def reports_callback(update, context):
    """مدیریت دکمه‌های منوی گزارش‌ها"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_allowed(user_id):
        return

    action = query.data

    # ==========================================
    # گزارش بر اساس دسته‌بندی
    # ==========================================
    if action == "report_category":

        await category_report_show_categories(
            update,
            context,
            from_callback=True
        )


    # ==========================================
    # هزینه‌های اخیر
    # ==========================================
    if action == "report_recent":

        context.user_data.clear()

        rows = get_recent_expenses(
            user_id,
            limit=10
        )

        if not rows:

            await query.edit_message_text(
                "📋 هنوز هیچ هزینه‌ای ثبت نشده."
            )

            return

        text = "📋 لیست هزینه‌ها\n\n"

        for display_number, (expense_id, amount, description, category, created_at) in enumerate(rows, start=1):

            date_part = to_jalali(created_at)

            time = (
                 created_at[11:16]
                if len(created_at) >= 16
                else ""
            )

            text += (
                f"#{display_number} {category}\n"
                f"💰 {amount:,} تومان\n"
                f"📝 {description}\n"
                f"📅 {date_part} | 🕐 {time}\n\n"
            )

        buttons = [[
            InlineKeyboardButton(
                "🔙 بازگشت به گزارش‌ها",
                callback_data="reports_menu"
            )
        ]]

        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )

        return

    # ==========================================
    # آمار کلی
    # ==========================================
    if action == "report_stats":

        context.user_data.clear()

        response = (
            supabase
            .table("expenses")
            .select("*")
            .eq("user_id", user_id)
            .execute()
        )

        rows = response.data or []

        if not rows:

            await query.edit_message_text(
                "📊 هنوز هیچ هزینه‌ای ثبت نشده."
            )

            return

        total = sum(
            int(row["amount"])
            for row in rows
        )

        count = len(rows)

        average = (
            total // count
            if count
            else 0
        )

        maximum = max(
            int(row["amount"])
            for row in rows
        )

        minimum = min(
            int(row["amount"])
            for row in rows
        )

        # امروز
        today = datetime.now(
            TEHRAN_TZ
        ).strftime("%Y-%m-%d")

        today_rows = [
            row
            for row in rows
            if str(row["created_at"]).startswith(today)
        ]

        today_count = len(today_rows)

        today_total = sum(
            int(row["amount"])
            for row in today_rows
        )

        # ماه جاری
        month = today[:7]

        month_rows = [
            row
            for row in rows
            if str(row["created_at"]).startswith(month)
        ]

        month_count = len(month_rows)

        month_total = sum(
            int(row["amount"])
            for row in month_rows
        )
        # تفکیک هزینه‌ها بر اساس دسته‌بندی
        category_totals = {}

        for row in rows:
            category = row.get("category") or "📦 سایر"
            amount = int(row["amount"])

            if category not in category_totals:
                category_totals[category] = 0

            category_totals[category] += amount

        category_text = ""

        if category_totals:
            category_text = "\n━━━━━━━━━━━━\n\n"
            category_text += "📂 هزینه‌ها بر اساس دسته‌بندی\n\n"

            sorted_categories = sorted(
                category_totals.items(),
                key=lambda x: x[1],
                reverse=True
            )

            for category, amount in sorted_categories:
                percentage = (
                    (amount / total) * 100
                    if total > 0
                    else 0
                )

                category_text += (
                    f"{category}\n"
                    f"💰 {amount:,} تومان "
                    f"({percentage:.1f}٪)\n\n"
                )
        today_jalali = to_jalali(today)

        text = "📊 خلاصه مالی\n\n"

        text += "💰 وضعیت کلی\n"
        text += f"مجموع هزینه‌ها: {total:,} تومان\n"
        text += f"تعداد هزینه‌ها: {count} مورد\n"
        text += f"میانگین هر هزینه: {average:,} تومان\n"
        text += f"🔺 بیشترین: {maximum:,} تومان\n"
        text += f"🔻 کمترین: {minimum:,} تومان\n\n"
        

        text += "━━━━━━━━━━━━\n\n"

        text += f"📅 امروز — {today_jalali}\n"
        text += f"🧾 {today_count} هزینه\n"
        text += f"💰 {today_total:,} تومان\n\n"

        text += "📅 این ماه\n"
        text += f"🧾 {month_count} هزینه\n"
        text += f"💰 {month_total:,} تومان\n"

        text += category_text

        buttons = [[
            InlineKeyboardButton(
                "🔙 بازگشت به گزارش‌ها",
                callback_data="reports_menu"
            )
        ]]

        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )

        return

    # ==========================================
    # گزارش پیشرفته
    # ==========================================
    if action == "report_advanced":

        context.user_data.clear()

        context.user_data["waiting_advanced_start"] = True

        keyboard = [
            [
                InlineKeyboardButton(
                    "📅 امروز",
                    callback_data="adv_today"
                ),
                InlineKeyboardButton(
                    "📅 این هفته",
                    callback_data="adv_this_week"
                )
            ],
            [
                InlineKeyboardButton(
                    "📅 هفته گذشته",
                    callback_data="adv_week"
                ),
                InlineKeyboardButton(
                    "📅 ماه جاری",
                    callback_data="adv_month"
                )
            ],
            [
                InlineKeyboardButton(
                    "📅 سه ماه اخیر",
                    callback_data="adv_quarter"
                ),
                InlineKeyboardButton(
                    "✏️ وارد کردن دستی",
                    callback_data="adv_manual"
                )
            ],
            [
                InlineKeyboardButton(
                    "🔙 بازگشت به گزارش‌ها",
                    callback_data="reports_menu"
                )
            ]
        ]

        await query.edit_message_text(
            "📅 گزارش بر اساس تاریخ\n\n"
            "یک بازه زمانی را انتخاب کن:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return

async def handle_message(update, context):
    user_id = update.effective_user.id
    if not is_allowed(user_id):
        await update.message.reply_text("⛔ شما اجازه استفاده از این ربات را ندارید.")
        return
    
    message = update.message.text.strip()
    
    if message == "🔙 بازگشت":
        await go_back(update, context)
        return
    # ==========================================
    # ویرایش کلمه کلیدی
    # ==========================================
    if context.user_data.get("keyword_edit_id"):
        keyword_id = context.user_data["keyword_edit_id"]

        if not message:
            return

        # پیدا کردن دسته‌بندی کلمه فعلی
        try:
            keyword_info = (
                supabase
                .table("category_keywords")
                .select("category_id")
                .eq("id", keyword_id)
                .maybe_single()
                .execute()
            )
            category_id = keyword_info.data.get("category_id") if keyword_info.data else None
        except Exception:
            logger.exception(f"خطا در یافتن کلمه کلیدی | user={user_id}")
            category_id = None

        if not category_id:
            context.user_data.clear()
            await update.message.reply_text(
                "❌ کلمه کلیدی پیدا نشد.",
                reply_markup=main_keyboard()
            )
            return

        # بررسی تکراری نبودن کلمه در همان دسته
        existing_keywords = (
            supabase
            .table("category_keywords")
            .select("id, keyword")
            .eq("category_id", category_id)
            .execute()
        )

        new_keyword = message.strip().casefold()

        for row in existing_keywords.data or []:
            if row["id"] != keyword_id:
                existing_keyword = (row.get("keyword") or "").strip().casefold()

                if existing_keyword == new_keyword:
                    context.user_data.clear()

                    await update.message.reply_text(
                        "❌ این کلمه قبلاً برای این دسته ثبت شده.",
                        reply_markup=main_keyboard()
                    )
                    return

        # انجام ویرایش
        response = (
            supabase
            .table("category_keywords")
            .update({"keyword": message})
            .eq("id", keyword_id)
            .execute()
        )

        if response.data:
            context.user_data.clear()

            await update.message.reply_text(
                f"✅ کلمه به «{message}» تغییر کرد."
            )

            text = await build_keywords_menu_text()

            await update.message.reply_text(
                text,
                reply_markup=keywords_menu_markup()
            )
        else:
            await update.message.reply_text(
                "❌ ویرایش کلمه انجام نشد.",
                reply_markup=back_keyboard()
            )

        return
        # ==========================================
    # افزودن کلمه کلیدی دسته‌بندی
    # ==========================================
    if context.user_data.get("keyword_add_category_id"):
        category_id = context.user_data["keyword_add_category_id"]

        if not message:
            return

        # بررسی تکراری نبودن کلمه
        response = (
            supabase
            .table("category_keywords")
            .select("id")
            .eq("keyword", message)
            .eq("category_id", category_id)
            .execute()
        )

        if response.data:
            await update.message.reply_text(
                "❌ این کلمه قبلاً برای این دسته ثبت شده.",
                reply_markup=back_keyboard()
            )
            return

        try:
            supabase.table("category_keywords").insert({
                "keyword": message,
                "category_id": category_id
            }).execute()

            context.user_data.clear()

            await update.message.reply_text(
                f"✅ کلمه «{message}» با موفقیت اضافه شد."
            )

            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="🔑 مدیریت کلمات دسته‌بندی",
                reply_markup=keywords_menu_markup()
            )

        except Exception:
            logger.exception(
                f"خطا در افزودن کلمه کلیدی | user={user_id}"
            )

            await update.message.reply_text(
                "❌ خطایی در افزودن کلمه کلیدی رخ داد.",
                reply_markup=back_keyboard()
            )

        return
    # ==========================================
    # مدیریت دسته‌بندی‌ها
    # ==========================================
    if context.user_data.get("waiting_category_add"):
        if not message:
            return
        if category_exists(message):
            await update.message.reply_text("❌ این دسته از قبل وجود دارد.", reply_markup=back_keyboard())
            return
        if add_category(message):
            context.user_data.clear()
            await update.message.reply_text(f"✅ دسته «{message}» اضافه شد.", reply_markup=main_keyboard())
        else:
            await update.message.reply_text("❌ خطا در افزودن دسته.", reply_markup=back_keyboard())
        return
    
    if context.user_data.get("waiting_category_rename"):
        category_id = context.user_data["rename_category_id"]
        if category_exists(message):
            await update.message.reply_text("❌ این نام قبلاً وجود دارد.", reply_markup=back_keyboard())
            return
        if rename_category(category_id, message):
            context.user_data.clear()
            await update.message.reply_text(f"✅ دسته به «{message}» تغییر کرد.", reply_markup=main_keyboard())
        else:
            await update.message.reply_text("❌ تغییر نام انجام نشد.", reply_markup=back_keyboard())
        return
        # ==========================================
    # گزارش دسته‌بندی - تاریخ شروع
    # ==========================================
    if context.user_data.get("waiting_category_report_start"):

        date_info = get_date_info(message)

        if not date_info:

            await update.message.reply_text(
                "❌ تاریخ نامعتبر است.\n\n"
                "مثال شمسی:\n"
                "1405-05-01\n\n"
                "مثال میلادی:\n"
                "2026-07-23",
                reply_markup=back_keyboard()
            )

            return

        start_date = date_info["gregorian"]

        context.user_data.pop(
            "waiting_category_report_start",
            None
        )

        context.user_data[
            "waiting_category_report_end"
        ] = True

        context.user_data[
            "category_report_start"
        ] = start_date

        await update.message.reply_text(
            "📅 تاریخ پایان را وارد کن:\n\n"
            "شمسی:\n"
            "1405-05-31\n\n"
            "میلادی:\n"
            "2026-08-22",
            reply_markup=back_keyboard()
        )

        return


    # ==========================================
    # گزارش دسته‌بندی - تاریخ پایان
    # ==========================================
    if context.user_data.get("waiting_category_report_end"):

        date_info = get_date_info(message)

        if not date_info:

            await update.message.reply_text(
                "❌ تاریخ نامعتبر است.\n\n"
                "مثال شمسی:\n"
                "1405-05-31\n\n"
                "مثال میلادی:\n"
                "2026-08-22",
                reply_markup=back_keyboard()
            )

            return

        end_date = date_info["gregorian"]

        start_date = context.user_data.get(
            "category_report_start"
        )

        if not start_date:

            context.user_data.clear()

            await update.message.reply_text(
                "❌ تاریخ شروع گزارش پیدا نشد.\n"
                "لطفاً گزارش را دوباره اجرا کن.",
                reply_markup=main_keyboard()
            )

            return

        if end_date < start_date:

            await update.message.reply_text(
                "❌ تاریخ پایان نمی‌تواند "
                "قبل از تاریخ شروع باشد.",
                reply_markup=back_keyboard()
            )

            return

        context.user_data.pop(
            "waiting_category_report_end",
            None
        )

        await show_category_report(
            update,
            context,
            start_date,
            end_date,
            "بازه دلخواه",
            page=0
        )

        return
    if context.user_data.get("waiting_advanced_start"):
        date_info = get_date_info(message)

        if not date_info:
            await update.message.reply_text(
                "❌ تاریخ نامعتبر است.\n\n"
                "مثال شمسی:\n"
                "1405-05-01\n\n"
                "مثال میلادی:\n"
                "2026-08-01",
                reply_markup=back_keyboard()
            )
            return

        start_date = date_info["gregorian"]

        context.user_data.clear()
        context.user_data["waiting_advanced_end"] = True
        context.user_data["advanced_start"] = start_date

        await update.message.reply_text(
            "📅 تاریخ پایان را وارد کن:\n\n"
            "شمسی:\n"
            "1405-05-23\n\n"
            "میلادی:\n"
            "2026-08-14",
            reply_markup=back_keyboard()
        )

        return
    
    if context.user_data.get("waiting_advanced_end"):
        date_info = get_date_info(message)

        if not date_info:
            await update.message.reply_text(
                "❌ تاریخ نامعتبر است.\n\n"
                "مثال شمسی:\n"
                "1405-05-23\n\n"
                "مثال میلادی:\n"
                "2026-08-14",
                reply_markup=back_keyboard()
            )
            return

        end_date = date_info["gregorian"]
        start_date = context.user_data["advanced_start"]

        if end_date < start_date:
            await update.message.reply_text(
                "❌ تاریخ پایان نمی‌تواند قبل از تاریخ شروع باشد.",
                reply_markup=back_keyboard()
            )
            return

        await show_advanced_report(
            update,
            context,
            start_date,
            end_date
        )

        return
    
    # ==========================================
    # منوی اصلی
    # ==========================================
    if message == "📥 ثبت هزینه":
        await expense_button(update, context)
        return
    
    if message == "⚡️ هزینه‌های سریع":
        await quick_expenses_menu(update, context)
        return
    
    if message == "📊 گزارش‌ها":
        await reports_menu(update, context)
        return
    
    if message == "✏️ مدیریت هزینه‌ها":
        await edit_delete_menu(update, context)
        return
    
    if message == "⚙️ تنظیمات":
        await settings(update, context)
        return
    
    # ==========================================
    # ثبت هزینه با انتخاب دسته
    # ==========================================
    if context.user_data.get("waiting_for_expense"):
        categories = [name for _, name in get_categories()]
        if message in categories:
            await choose_category(update, context, message)
            return

        # ==========================================
    # ثبت مبلغ و توضیح بعد از انتخاب دسته
    # ==========================================
    if context.user_data.get("waiting_for_amount"):
        parsed = parse_expense_text(message)

        if not parsed:
            await update.message.reply_text(
                "❌ فرمت درست نیست.\n\n"
                "مثال:\n"
                "85000 ناهار",
                reply_markup=back_keyboard()
            )
            return

        amount, description = parsed
        category = context.user_data.get("selected_category", "📦 سایر")

        add_expense(user_id, amount, description, category)

        context.user_data.clear()

        await update.message.reply_text(
            f"✅ هزینه ثبت شد!\n\n"
            f"{category}\n"
            f"💰 {amount:,} تومان\n"
            f"📝 {description}",
            reply_markup=main_keyboard()
        )
        return
    # ==========================================
    # ویرایش هزینه
    # ==========================================
    if context.user_data.get("waiting_for_edit"):
        parsed = parse_expense_text(message)
        if not parsed:
            await update.message.reply_text("❌ فرمت درست نیست.\n\nمثال:\n95000 ناهار رستوران", reply_markup=back_keyboard())
            return
        amount, description = parsed
        expense_id = context.user_data["editing_expense"]
        category = context.user_data.get("editing_category", "📦 سایر")
        updated = update_expense(user_id, expense_id, amount, description, category)
        context.user_data.clear()
        if updated:
            await update.message.reply_text(
                f"✅ هزینه #{expense_id} ویرایش شد.\n\n{category}\n💰 {amount:,} تومان\n📝 {description}",
                reply_markup=main_keyboard()
            )
        return
    

        # ==========================================
    # افزودن هزینه سریع جدید (روش خیلی ساده)
    # ==========================================
    if context.user_data.get("waiting_quick_add"):
        category = context.user_data.get(
            "quick_add_category",
            "📦 سایر"
        )

        message_text = normalize_digits(message.strip())

        # حالت:
        # 85000 ناهار
        parsed = parse_expense_text(message_text)

        if parsed:
            amount, name = parsed

        else:
            # حالت:
            # 85000
            amount = parse_amount(message_text)

            if amount is None:
                await update.message.reply_text(
                    "❌ مبلغ نامعتبر است.\n\n"
                    "مثال:\n"
                    "85000\n\n"
                    "یا:\n"
                    "85000 ناهار",
                    reply_markup=back_keyboard()
                )
                return

            name = category

        add_quick_expense(
            user_id,
            name,
            amount,
            category
        )

        context.user_data.clear()

        await update.message.reply_text(
            "✅ هزینه سریع با موفقیت اضافه شد!\n\n"
            f"📝 {name}\n"
            f"💰 {amount:,} تومان\n"
            f"📂 {category}",
            reply_markup=main_keyboard()
        )

        return
    # ==========================================
    # ✅ ویرایش هزینه سریع
    # ==========================================
    if context.user_data.get("waiting_quick_edit"):
        quick_id = context.user_data.get("quick_edit_id")
    
        if not quick_id:
            context.user_data.clear()
            await update.message.reply_text(
                "❌ اطلاعات ویرایش پیدا نشد.",
                reply_markup=main_keyboard()
            )
            return
    
        message_text = normalize_digits(message.strip())
    
        # دریافت اطلاعات فعلی
        response = (
            supabase
            .table("quick_expenses")
            .select("*")
            .eq("id", quick_id)
            .eq("user_id", user_id)
            .execute()
        )
    
        if not response.data:
            context.user_data.clear()
            await update.message.reply_text(
                "❌ هزینه سریع پیدا نشد.",
                reply_markup=main_keyboard()
            )
            return
    
        item = response.data[0]
    
        # مقادیر پیش‌فرض = اطلاعات فعلی
        name = item["name"]
        amount = item["amount"]
        category = item["category"]
    
        # ==========================================
        # حالت کامل:
        # نام|مبلغ|دسته‌بندی
        # ==========================================
        if "|" in message_text:
    
            parts = [part.strip() for part in message_text.split("|")]
    
            if len(parts) != 3:
                await update.message.reply_text(
                    "❌ فرمت اشتباه است.\n\n"
                    "فرمت صحیح:\n"
                    "نام|مبلغ|دسته‌بندی\n\n"
                    "مثال:\n"
                    "صبحانه بیرون|80000|🍔 غذا",
                    reply_markup=back_keyboard()
                )
                return
    
            new_name = parts[0]
            new_amount = parse_amount(parts[1])
            new_category = parts[2]
    
            if not new_name:
                await update.message.reply_text(
                    "❌ نام هزینه نمی‌تواند خالی باشد.",
                    reply_markup=back_keyboard()
                )
                return
    
            if new_amount is None:
                await update.message.reply_text(
                    "❌ مبلغ نامعتبر است.",
                    reply_markup=back_keyboard()
                )
                return
    
            categories = get_categories()
            category_names = [cat_name for _, cat_name in categories]
    
            if new_category not in category_names:
                await update.message.reply_text(
                    f"❌ دسته‌بندی «{new_category}» وجود ندارد.",
                    reply_markup=back_keyboard()
                )
                return
    
            # استفاده از مقادیر جدید
            name = new_name
            amount = new_amount
            category = new_category
    
        # ==========================================
        # فقط مبلغ
        # ==========================================
        else:
    
            new_amount = parse_amount(message_text)
    
            if new_amount is None:
                await update.message.reply_text(
                    "❌ مبلغ باید یک عدد مثبت باشد.\n\n"
                    "مثال:\n"
                    "75000\n\n"
                    "یا برای تغییر کامل:\n"
                    "صبحانه بیرون|80000|🍔 غذا",
                    reply_markup=back_keyboard()
                )
                return
    
            # فقط مبلغ تغییر می‌کند
            amount = new_amount
    
        # ==========================================
        # ذخیره تغییرات
        # ==========================================
        updated = update_quick_expense(
            user_id,
            quick_id,
            name,
            amount,
            category
        )
    
        context.user_data.clear()
    
        if updated:
            await update.message.reply_text(
                "✅ هزینه سریع ویرایش شد!\n\n"
                f"📝 {name}\n"
                f"💰 {amount:,} تومان\n"
                f"📂 {category}",
                reply_markup=main_keyboard()
            )
        else:
            await update.message.reply_text(
                "❌ ویرایش هزینه سریع انجام نشد.",
                reply_markup=main_keyboard()
            )
    
        return

    # ==========================================
    # ثبت سریع هزینه (بدون دسته)
    # ==========================================
    parsed = parse_expense_text(message)
    if parsed:
        amount, description = parsed
        category = detect_category(description)
        add_expense(user_id, amount, description, category)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ هزینه ثبت شد!\n\n{category}\n💰 {amount:,} تومان\n📝 {description}",
            reply_markup=main_keyboard()
        )
        return
    
    await update.message.reply_text(
        "❓ از دکمه‌های منو استفاده کن.\n\nیا برای ثبت سریع بنویس:\n85 ناهار",
        reply_markup=main_keyboard()
    )

async def reports_menu_callback(update, context):
    """منوی اصلی گزارش‌ها"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_allowed(user_id):
        return

    context.user_data.clear()

    await query.edit_message_text(
        REPORTS_MENU_TEXT,
        reply_markup=reports_menu_markup()
    )


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    """ثبت خطاهای پیش‌بینی‌نشده و اطلاع‌رسانی به کاربر"""
    logger.error("خطای پردازش آپدیت:", exc_info=context.error)
    try:
        if isinstance(update, Update) and update.effective_user:
            await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="❌ خطایی رخ داد. لطفاً دوباره تلاش کن."
            )
    except Exception:
        pass


# ==========================================
# Health Server برای Render
# ==========================================

class HealthHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        if self.path == "/health":
            self.send_response(200)
            self.send_header("Content-type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write(b"OK")
        else:
            self.send_response(404)
            self.end_headers()

    def do_HEAD(self):
        if self.path == "/health":
            self.send_response(200)
            self.send_header("Content-type", "text/plain; charset=utf-8")
            self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, format, *args):
        return


def run_health_server():
    port = int(os.environ.get("PORT", 10000))
    server = HTTPServer(("0.0.0.0", port), HealthHandler)
    print(f"Health server running on port {port}")
    server.serve_forever()



def main():

    # آماده‌سازی دیتابیس (دسته‌ها و کلمات کلیدی پیش‌فرض)
    init_db()

    # سرور Health برای Render
    health_thread = threading.Thread(
        target=run_health_server,
        daemon=True
    )
    health_thread.start()
    
    request = HTTPXRequest(connect_timeout=60, read_timeout=60, write_timeout=60, pool_timeout=60)
    app = Application.builder().token(TOKEN).request(request).get_updates_request(request).build()
    app.add_handler(
    CallbackQueryHandler(
        advanced_report_page_callback,
        pattern=r"^advanced_report_page:"
    )
)
    app.add_handler(CommandHandler("start", start))
    
    app.add_error_handler(error_handler)

    app.add_handler(CallbackQueryHandler(delete_callback, pattern=r"^delete:\d+$"))
    app.add_handler(CallbackQueryHandler(confirm_delete_callback, pattern=r"^confirm_delete:\d+$"))
    app.add_handler(CallbackQueryHandler(cancel_delete_callback, pattern=r"^cancel_delete$"))
    app.add_handler(CallbackQueryHandler(edit_callback, pattern=r"^edit:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_categories, pattern=r"^manage_categories$"))
    app.add_handler(CallbackQueryHandler(manage_keywords, pattern=r"^manage_keywords$"))
    app.add_handler(CallbackQueryHandler(category_add_callback, pattern=r"^category_add$"))
    app.add_handler(CallbackQueryHandler(category_rename_callback, pattern=r"^category_rename$"))
    app.add_handler(CallbackQueryHandler(rename_select_callback, pattern=r"^rename_select:\d+$"))
    app.add_handler(CallbackQueryHandler(category_delete_callback, pattern=r"^category_delete$"))
    app.add_handler(CallbackQueryHandler(delete_category_callback, pattern=r"^delete_category:\d+$"))
    app.add_handler(CallbackQueryHandler(confirm_category_delete_callback, pattern=r"^confirm_category_delete:\d+$"))
    app.add_handler(CallbackQueryHandler(settings_menu_callback, pattern=r"^settings_menu$"))
    app.add_handler(CallbackQueryHandler(back_callback, pattern=r"^back_menu$"))
    app.add_handler(CallbackQueryHandler(edit_page_callback, pattern=r"^edit_page:\d+$"))
    app.add_handler(CallbackQueryHandler(advanced_quick_callback, pattern=r"^adv_"))
    app.add_handler(CallbackQueryHandler(ignore_callback, pattern=r"^ignore$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CallbackQueryHandler(keyword_add_callback, pattern=r"^keyword_add$"))
    app.add_handler(CallbackQueryHandler(
        keyword_add_category_callback,
        pattern=r"^keyword_add_cat:\d+$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_edit_callback,
    pattern=r"^keyword_edit$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_edit_category_callback,
    pattern=r"^keyword_edit_category:\d+$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_edit_select_callback,
    pattern=r"^keyword_edit_select:\d+$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_delete_callback,
    pattern=r"^keyword_delete$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_delete_category_callback,
    pattern=r"^keyword_delete_category:\d+$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_delete_select_callback,
    pattern=r"^keyword_delete_select:\d+$"
    ))
    app.add_handler(CallbackQueryHandler(
    keyword_delete_confirm_callback,
    pattern=r"^keyword_delete_confirm:\d+$"
    ))
    app.add_handler(
    CallbackQueryHandler(
        export_excel_callback,
        pattern=r"^report_excel$"
    )
    )
    # ==========================================
    # مدیریت هزینه‌های سریع
    # ==========================================
    app.add_handler(CallbackQueryHandler(quick_manage_callback, pattern=r"^quick_manage$"))
    app.add_handler(CallbackQueryHandler(quick_add_callback, pattern=r"^quick_add$"))
    app.add_handler(CallbackQueryHandler(quick_edit_callback, pattern=r"^quick_edit$"))
    app.add_handler(CallbackQueryHandler(quick_delete_callback, pattern=r"^quick_delete$"))
    app.add_handler(CallbackQueryHandler(quick_delete_ask_callback, pattern=r"^quick_delete_ask_\d+$"))
    app.add_handler(CallbackQueryHandler(quick_menu_callback, pattern=r"^quick_menu$"))
    app.add_handler(CallbackQueryHandler(quick_callback, pattern=r"^quick_\d+$"))
    app.add_handler(CallbackQueryHandler(quick_edit_select_callback, pattern=r"^quick_edit_select_"))
    app.add_handler(CallbackQueryHandler(quick_delete_confirm_callback, pattern=r"^quick_delete_confirm_"))
    app.add_handler(CallbackQueryHandler(quick_add_category_callback, pattern=r"^quick_add_cat_"))
    app.add_handler(CallbackQueryHandler(category_report_callback, pattern=r"^cat_report_"))
    app.add_handler(CallbackQueryHandler(reports_menu_callback, pattern=r"^reports_menu$"))
    app.add_handler(CallbackQueryHandler(reports_callback, pattern=r"^report_(advanced|recent|stats|category)$"))
   
    print("✅ ربات اجرا شد!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
